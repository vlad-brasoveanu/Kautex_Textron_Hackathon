import io
import os
import csv
import re
import uuid
import datetime
import hashlib
import difflib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import urllib.request
import urllib.error
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Header, status
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import models, schemas
from database import engine, Base, get_db

# Create DB Tables
Base.metadata.create_all(bind=engine)

# Directory where raw copies of every CSV/Excel upload are kept, so a past
# upload can be listed in "Upload History" and re-applied later without the
# user needing to keep the original file around.
UPLOAD_STORAGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_STORAGE_DIR, exist_ok=True)

# Lightweight migration: create_all only creates missing tables, it does not add
# columns to tables that already existed before these fields were introduced.
def migrate_add_missing_columns():
    with engine.connect() as conn:
        existing_cols = {row[1] for row in conn.exec_driver_sql("PRAGMA table_info(users)").fetchall()}
        for col in ("email", "department", "position", "supervisor"):
            if col not in existing_cols:
                conn.exec_driver_sql(f"ALTER TABLE users ADD COLUMN {col} VARCHAR")

        for table in ("employees", "topics"):
            existing_cols = {row[1] for row in conn.exec_driver_sql(f"PRAGMA table_info({table})").fetchall()}
            if "is_deleted" not in existing_cols:
                conn.exec_driver_sql(f"ALTER TABLE {table} ADD COLUMN is_deleted BOOLEAN DEFAULT 0")
            if "deleted_at" not in existing_cols:
                conn.exec_driver_sql(f"ALTER TABLE {table} ADD COLUMN deleted_at DATETIME")

        conn.commit()

migrate_add_missing_columns()

app = FastAPI(title="Digital Engineering Planning Dashboard API")

# Password hashing helper
def hash_password(password: str) -> str:
    salt = "textron_hackathon_salt_2026"
    return hashlib.sha256((password + salt).encode('utf-8')).hexdigest()

# HTTP Bearer Auth Guard
security = HTTPBearer()

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    token = credentials.credentials
    if not token or not token.startswith("token_"):
        raise HTTPException(status_code=401, detail="Invalid or missing session token")
        
    parts = token.split("_")
    if len(parts) < 3:
        raise HTTPException(status_code=401, detail="Malformed session token")
        
    username = parts[1]
    role = "_".join(parts[2:])
    
    # Verify in DB
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or user.role != role:
        raise HTTPException(status_code=401, detail="User session context invalid")
        
    return user

def require_admin(user: models.User = Depends(get_current_user)):
    if user.role not in ["admin", "master_admin"]:
        raise HTTPException(status_code=403, detail="Admin privileges required for this action")
    return user


def require_master(user: models.User = Depends(get_current_user)):
    if user.role != "master_admin":
        raise HTTPException(status_code=403, detail="Master Admin privileges required for this action")
    return user


def write_system_log(db: Session, username: str, action: str, details: str):
    try:
        log = models.SystemLog(username=username, action=action, details=details)
        db.add(log)
        db.commit()
    except Exception as e:
        print("Failed to write system audit log:", e)


# Auth Login Endpoint
@app.post("/api/auth/login", response_model=schemas.UserResponse)
def login(payload: schemas.UserLogin, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == payload.username).first()
    if not user or user.password_hash != hash_password(payload.password):
        write_system_log(db, username=payload.username, action="Failed Login", details="Invalid username or password")
        raise HTTPException(status_code=401, detail="Invalid username or password")
        
    write_system_log(db, username=user.username, action="Login", details=f"Successful login. Role: {user.role}")
    token = f"token_{user.username}_{user.role}"
    return {
        "username": user.username,
        "role": user.role,
        "access_token": token,
        "name": user.name
    }


# Auth Register Endpoint
@app.post("/api/auth/register", response_model=schemas.UserResponse)
def register(payload: schemas.UserRegister, db: Session = Depends(get_db)):
    existing = db.query(models.User).filter(models.User.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
        
    new_user = models.User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        name=payload.name,
        role=payload.role or "user"
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    write_system_log(db, username=new_user.username, action="Registration", details=f"New user registered with role: {new_user.role}")
    token = f"token_{new_user.username}_{new_user.role}"
    return {
        "username": new_user.username,
        "role": new_user.role,
        "access_token": token,
        "name": new_user.name
    }


# Get all users (Admin/Master Admin only)
@app.get("/api/users", response_model=List[schemas.UserManageResponse])
def list_users(db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    return db.query(models.User).all()


# Create User (Admin/Master Admin only, subject to role hierarchy)
@app.post("/api/users", response_model=schemas.UserManageResponse)
def create_user(payload: schemas.UserRegister, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    existing = db.query(models.User).filter(models.User.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
        
    target_role = payload.role or "user"
    
    # Role validation
    if current_user.role == "admin":
        if target_role != "user":
            raise HTTPException(status_code=403, detail="Admins are only authorized to create standard users")
    elif current_user.role == "master_admin":
        if target_role not in ["admin", "user"]:
            raise HTTPException(status_code=403, detail="Cannot create master admin accounts")
    else:
        raise HTTPException(status_code=403, detail="Unauthorized role creation")
        
    new_user = models.User(
        username=payload.username,
        password_hash=hash_password(payload.password),
        name=payload.name,
        role=target_role,
        email=payload.email or None,
        department=payload.department or None,
        position=payload.position or None,
        supervisor=payload.supervisor or None
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    write_system_log(
        db,
        username=current_user.username,
        action="Registration",
        details=f"Created user '{new_user.username}' with role '{new_user.role}' and name '{new_user.name}'"
    )
    return new_user


# Update User (Master Admin can edit any account; Admin can edit Admin/User
# accounts but not Master Admin; only Master Admin can change roles)
@app.put("/api/users/{user_id}", response_model=schemas.UserManageResponse)
def update_user(user_id: int, payload: schemas.UserUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User account not found")

    if target_user.role == "master_admin" and current_user.role != "master_admin":
        raise HTTPException(status_code=403, detail="Only Master Admin can edit Master Admin accounts")

    if payload.role is not None:
        if current_user.role != "master_admin":
            raise HTTPException(status_code=403, detail="Only Master Admin can change user roles")
        if target_user.role == "master_admin":
            raise HTTPException(status_code=400, detail="Cannot change the Master Admin's own role")
        if payload.role not in ("admin", "user"):
            raise HTTPException(status_code=400, detail="Role must be 'admin' or 'user'")
        target_user.role = payload.role

    if payload.name is not None:
        target_user.name = payload.name
    if payload.email is not None:
        target_user.email = payload.email or None
    if payload.department is not None:
        target_user.department = payload.department or None
    if payload.position is not None:
        target_user.position = payload.position or None
    if payload.supervisor is not None:
        target_user.supervisor = payload.supervisor or None
    if payload.password:
        target_user.password_hash = hash_password(payload.password)

    db.commit()
    db.refresh(target_user)

    write_system_log(
        db,
        username=current_user.username,
        action="Edit User",
        details=f"Updated user account '{target_user.username}' (Name: '{target_user.name}', Role: '{target_user.role}')"
    )
    return target_user


# Delete User (Admin/Master Admin only, subject to deletion safeguards)
@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User account not found")
        
    if target_user.role == "master_admin":
        raise HTTPException(status_code=400, detail="Master admin accounts are protected and cannot be deleted")
        
    if target_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own active session account")
        
    if current_user.role == "admin":
        if target_user.role != "user":
            raise HTTPException(status_code=403, detail="Admins can only delete regular user accounts")
            
    # Delete user
    db.delete(target_user)
    db.commit()
    
    write_system_log(
        db,
        username=current_user.username,
        action="User Deletion",
        details=f"Deleted user '{target_user.username}' (Name: '{target_user.name}', Role: '{target_user.role}')"
    )
    return {"status": "success", "message": f"Successfully deleted user '{target_user.username}'"}


class ExportLogPayload(BaseModel):
    report_name: str
    format: str


@app.post("/api/reports/log-export")
def log_report_export(payload: ExportLogPayload, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    write_system_log(
        db,
        username=current_user.username,
        action="Export Report",
        details=f"Exported report: {payload.report_name} in format: {payload.format}"
    )
    return {"status": "success"}


@app.get("/api/admin/logs", response_model=List[schemas.AuditLogResponse])
def get_admin_logs(db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    logs = db.query(models.SystemLog).order_by(models.SystemLog.timestamp.desc()).all()
    return logs


@app.delete("/api/admin/logs")
def clear_admin_logs(db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    count = db.query(models.SystemLog).delete()
    db.commit()
    # Written after the wipe so it survives as the first entry of the fresh log.
    write_system_log(db, username=master.username, action="Clear Audit Logs", details=f"Permanently cleared {count} audit log entries")
    return {"message": f"Cleared {count} audit log entries"}


def get_planning_rag_context(db: Session, scenario_id: int) -> str:
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        return "No active planning scenario data."
        
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == scenario_id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == scenario_id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).filter(models.Employee.scenario_id == scenario_id, models.Employee.is_deleted == False).all()
    additional_costs = db.query(models.AdditionalCost).join(models.Topic).filter(models.Topic.scenario_id == scenario_id, models.Topic.is_deleted == False).all()

    lines = []
    lines.append(f"Scenario Name: {scenario.name}")
    lines.append(f"Scenario Description: {scenario.description or ''}")
    
    lines.append("\n## Employees")
    for e in employees:
        lines.append(f"- Name: {e.name}, Team: {e.team}, Location: {e.location}, Rate: ${e.hourly_rate}/hr, Hours: {e.available_hours}h/yr, Status: {e.status}")
        
    lines.append("\n## Projects / Topics")
    for t in topics:
        lines.append(f"- Topic Name: {t.name}, Category: {t.category}, Area: {t.area or 'N/A'}, Recovery Savings: ${t.recovery}")
        
    lines.append("\n## Allocations (%)")
    alloc_map = {(a.employee_id, a.topic_id): a for a in allocations}
    emp_names = {e.id: e.name for e in employees}
    top_names = {t.id: t.name for t in topics}
    for (eid, tid), a in alloc_map.items():
        if eid in emp_names and tid in top_names:
            lines.append(f"- Employee '{emp_names[eid]}' is allocated to Topic '{top_names[tid]}' at {a.percentage}%" + (f" (Reason/Comment: {a.comment})" if a.comment else ""))
            
    lines.append("\n## Additional Cost Items")
    for ac in additional_costs:
        if ac.topic_id in top_names:
            lines.append(f"- Topic '{top_names[ac.topic_id]}': Category '{ac.category}', Type: '{ac.cost_type}', Amount: ${ac.amount}")
            
    return "\n".join(lines)


def query_local_ollama(prompt: str) -> Optional[str]:
    url = "http://127.0.0.1:11434/api/generate"
    data = json.dumps({
        "model": "llama3",
        "prompt": prompt,
        "stream": False,
        "options": {
            "temperature": 0.2
        }
    }).encode("utf-8")
    
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=8.0) as response:
            resp_data = json.loads(response.read().decode("utf-8"))
            return resp_data.get("response")
    except Exception as e:
        print(f"Ollama local connection fallback active: {e}")
    return None

# Setup CORS for development convenience
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Helper function to get the current active scenario
def get_active_scenario_db(db: Session):
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()
    if not scenario:
        # Fallback to create one if somehow none is active
        scenario = models.Scenario(name="Default Scenario", description="Auto-created default scenario", is_active=True)
        db.add(scenario)
        db.commit()
        db.refresh(scenario)
    return scenario

# ==========================================
# 1. SCENARIOS API
# ==========================================

@app.get("/api/scenarios", response_model=List[schemas.ScenarioResponse])
def get_scenarios(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.Scenario).all()

@app.get("/api/scenarios/active", response_model=schemas.ScenarioResponse)
def get_active_scenario(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return get_active_scenario_db(db)

@app.post("/api/scenarios/active/{scenario_id}")
def set_active_scenario(scenario_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    db.query(models.Scenario).update({models.Scenario.is_active: False})
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        db.rollback()
        raise HTTPException(status_code=404, detail="Scenario not found")
    scenario.is_active = True
    db.commit()
    return {"message": f"Active scenario switched to: {scenario.name}", "scenario": scenario}

@app.post("/api/scenarios", response_model=schemas.ScenarioResponse)
def create_scenario(scenario: schemas.ScenarioCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    # Deactivate other scenarios
    db.query(models.Scenario).update({models.Scenario.is_active: False})
    
    db_scenario = models.Scenario(
        name=scenario.name,
        description=scenario.description,
        is_active=True
    )
    db.add(db_scenario)
    db.commit()
    db.refresh(db_scenario)
    return db_scenario

@app.post("/api/scenarios/{scenario_id}/clone", response_model=schemas.ScenarioResponse)
def clone_scenario(scenario_id: int, clone_data: schemas.ScenarioClone, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    source_scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not source_scenario:
        raise HTTPException(status_code=404, detail="Source scenario not found")
    
    # Deactivate others
    db.query(models.Scenario).update({models.Scenario.is_active: False})
    
    # Create new scenario
    new_scenario = models.Scenario(
        name=clone_data.new_name,
        description=clone_data.new_description or f"Clone of {source_scenario.name}",
        is_active=True
    )
    db.add(new_scenario)
    db.commit()
    db.refresh(new_scenario)
    
    # 1. Clone Employees (Trash excluded - a deleted employee shouldn't reappear in the clone)
    db_employees = db.query(models.Employee).filter(models.Employee.scenario_id == scenario_id, models.Employee.is_deleted == False).all()
    emp_id_mapping = {} # Old ID -> New ID
    for emp in db_employees:
        new_emp = models.Employee(
            scenario_id=new_scenario.id,
            name=emp.name,
            team=emp.team,
            department=emp.department,
            location=emp.location,
            available_hours=emp.available_hours,
            hourly_rate=emp.hourly_rate,
            status=emp.status,
            manager=emp.manager,
            notes=emp.notes
        )
        db.add(new_emp)
        db.commit()
        db.refresh(new_emp)
        emp_id_mapping[emp.id] = new_emp.id
        
    # 2. Clone Topics & Additional Costs (Trash excluded)
    db_topics = db.query(models.Topic).filter(models.Topic.scenario_id == scenario_id, models.Topic.is_deleted == False).all()
    topic_id_mapping = {} # Old ID -> New ID
    for topic in db_topics:
        new_topic = models.Topic(
            scenario_id=new_scenario.id,
            name=topic.name,
            category=topic.category,
            area=topic.area,
            description=topic.description,
            objective=topic.objective,
            deliverables=topic.deliverables,
            justification=topic.justification,
            status=topic.status,
            comments=topic.comments,
            notes=topic.notes,
            recovery=topic.recovery
        )
        db.add(new_topic)
        db.commit()
        db.refresh(new_topic)
        topic_id_mapping[topic.id] = new_topic.id
        
        # Clone Additional Costs for this topic
        for cost in topic.additional_costs:
            new_cost = models.AdditionalCost(
                topic_id=new_topic.id,
                cost_type=cost.cost_type,
                category=cost.category,
                amount=cost.amount,
                notes=cost.notes
            )
            db.add(new_cost)
        db.commit()
        
    # 3. Clone Allocations
    for old_emp_id, new_emp_id in emp_id_mapping.items():
        old_allocations = db.query(models.Allocation).filter(models.Allocation.employee_id == old_emp_id).all()
        for alloc in old_allocations:
            if alloc.topic_id in topic_id_mapping:
                new_alloc = models.Allocation(
                    employee_id=new_emp_id,
                    topic_id=topic_id_mapping[alloc.topic_id],
                    percentage=alloc.percentage,
                    comment=alloc.comment
                )
                db.add(new_alloc)
        db.commit()
        
    return new_scenario

@app.delete("/api/scenarios/{scenario_id}")
def delete_scenario(scenario_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    
    # If active was deleted, make another one active if exists
    was_active = scenario.is_active
    db.delete(scenario)
    db.commit()
    
    if was_active:
        next_scenario = db.query(models.Scenario).first()
        if next_scenario:
            next_scenario.is_active = True
            db.commit()
            
    return {"message": "Scenario deleted successfully"}
    

@app.get("/api/scenarios/{scenario_id}/backup")
def backup_scenario(scenario_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
        
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == scenario_id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == scenario_id, models.Topic.is_deleted == False).all()

    emp_ids = [e.id for e in employees]
    topic_ids = [t.id for t in topics]

    allocations = db.query(models.Allocation).filter(
        models.Allocation.employee_id.in_(emp_ids) if emp_ids else False
    ).all()
    
    additional_costs = db.query(models.AdditionalCost).filter(
        models.AdditionalCost.topic_id.in_(topic_ids) if topic_ids else False
    ).all()
    
    data = {
        "version": "1.0",
        "name": scenario.name,
        "description": scenario.description or "",
        "employees": [
            {
                "name": e.name,
                "team": e.team,
                "department": e.department,
                "location": e.location,
                "available_hours": e.available_hours,
                "hourly_rate": e.hourly_rate,
                "manager": e.manager,
                "status": e.status,
                "notes": e.notes
            } for e in employees
        ],
        "topics": [
            {
                "name": t.name,
                "category": t.category,
                "area": t.area,
                "description": t.description,
                "objective": t.objective,
                "deliverables": t.deliverables,
                "justification": t.justification,
                "status": t.status,
                "comments": t.comments,
                "notes": t.notes,
                "recovery": t.recovery
            } for t in topics
        ],
        "allocations": [
            {
                "employee_name": next((e.name for e in employees if e.id == a.employee_id), None),
                "topic_name": next((t.name for t in topics if t.id == a.topic_id), None),
                "percentage": a.percentage,
                "comment": a.comment
            } for a in allocations
        ],
        "additional_costs": [
            {
                "topic_name": next((t.name for t in topics if t.id == c.topic_id), None),
                "cost_type": c.cost_type,
                "category": c.category,
                "amount": c.amount,
                "notes": c.notes
            } for c in additional_costs
        ]
    }
    
    write_system_log(
        db,
        username=current_user.username,
        action="Export Report",
        details=f"Exported JSON backup for scenario '{scenario.name}'"
    )
    return data


@app.post("/api/scenarios/{scenario_id}/restore")
def restore_scenario(scenario_id: int, payload: schemas.RestorePayload, db: Session = Depends(get_db), current_user: models.User = Depends(require_admin)):
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
        
    db_employees = db.query(models.Employee).filter(models.Employee.scenario_id == scenario_id).all()
    db_topics = db.query(models.Topic).filter(models.Topic.scenario_id == scenario_id).all()
    
    emp_ids = [e.id for e in db_employees]
    topic_ids = [t.id for t in db_topics]
    
    if emp_ids:
        db.query(models.Allocation).filter(models.Allocation.employee_id.in_(emp_ids)).delete(synchronize_session=False)
    if topic_ids:
        db.query(models.AdditionalCost).filter(models.AdditionalCost.topic_id.in_(topic_ids)).delete(synchronize_session=False)
        
    db.query(models.Employee).filter(models.Employee.scenario_id == scenario_id).delete(synchronize_session=False)
    db.query(models.Topic).filter(models.Topic.scenario_id == scenario_id).delete(synchronize_session=False)
    db.commit()
    
    scenario.name = payload.name
    scenario.description = payload.description
    db.commit()
    
    emp_name_map = {}
    for emp_data in payload.employees:
        emp = models.Employee(
            name=emp_data["name"],
            team=emp_data.get("team", "General"),
            department=emp_data.get("department", "General"),
            location=emp_data.get("location", "US"),
            available_hours=emp_data.get("available_hours", 1800.0),
            hourly_rate=emp_data.get("hourly_rate", 50.0),
            manager=emp_data.get("manager"),
            status=emp_data.get("status", "Active"),
            notes=emp_data.get("notes"),
            scenario_id=scenario_id
        )
        db.add(emp)
        db.commit()
        db.refresh(emp)
        emp_name_map[emp.name] = emp.id
        
    topic_name_map = {}
    for topic_data in payload.topics:
        topic = models.Topic(
            name=topic_data["name"],
            category=topic_data.get("category", "General"),
            area=topic_data.get("area"),
            description=topic_data.get("description"),
            objective=topic_data.get("objective"),
            deliverables=topic_data.get("deliverables"),
            justification=topic_data.get("justification"),
            status=topic_data.get("status", "Active"),
            comments=topic_data.get("comments"),
            notes=topic_data.get("notes"),
            recovery=topic_data.get("recovery", 0.0),
            scenario_id=scenario_id
        )
        db.add(topic)
        db.commit()
        db.refresh(topic)
        topic_name_map[topic.name] = topic.id

    for alloc_data in payload.allocations:
        emp_id = emp_name_map.get(alloc_data["employee_name"])
        topic_id = topic_name_map.get(alloc_data["topic_name"])
        if emp_id and topic_id:
            alloc = models.Allocation(
                employee_id=emp_id,
                topic_id=topic_id,
                percentage=alloc_data["percentage"],
                comment=alloc_data.get("comment", "")
            )
            db.add(alloc)

    for cost_data in payload.additional_costs:
        topic_id = topic_name_map.get(cost_data["topic_name"])
        if topic_id:
            ac = models.AdditionalCost(
                topic_id=topic_id,
                cost_type=cost_data["cost_type"],
                category=cost_data.get("category", "Other"),
                amount=cost_data.get("amount", 0.0),
                notes=cost_data.get("notes")
            )
            db.add(ac)
            
    db.commit()
    write_system_log(
        db,
        username=current_user.username,
        action="Import CSV",
        details=f"Restored scenario '{scenario.name}' from JSON backup"
    )
    return {"status": "success", "message": f"Successfully restored scenario '{scenario.name}'"}

# ==========================================
# 2. EMPLOYEES API (CRUD + Scenario-Scoped)
# ==========================================

@app.get("/api/employees", response_model=List[schemas.EmployeeResponse])
def get_employees(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    return db.query(models.Employee).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False
    ).all()

@app.post("/api/employees", response_model=schemas.EmployeeResponse)
def create_employee(employee: schemas.EmployeeCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    db_employee = models.Employee(
        scenario_id=active_scenario.id,
        name=employee.name,
        team=employee.team,
        department=employee.department,
        location=employee.location,
        available_hours=employee.available_hours,
        hourly_rate=employee.hourly_rate,
        status=employee.status,
        manager=employee.manager,
        notes=employee.notes
    )
    db.add(db_employee)
    db.commit()
    db.refresh(db_employee)
    return db_employee

@app.put("/api/employees/{employee_id}", response_model=schemas.EmployeeResponse)
def update_employee(employee_id: int, employee: schemas.EmployeeCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    db_employee.name = employee.name
    db_employee.team = employee.team
    db_employee.department = employee.department
    db_employee.location = employee.location
    db_employee.available_hours = employee.available_hours
    db_employee.hourly_rate = employee.hourly_rate
    db_employee.status = employee.status
    db_employee.manager = employee.manager
    db_employee.notes = employee.notes
    
    db.commit()
    db.refresh(db_employee)
    return db_employee

@app.patch("/api/employees/bulk", response_model=List[schemas.EmployeeResponse])
def bulk_edit_employees(payload: schemas.BulkEmployeeEdit, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    db_employees = db.query(models.Employee).filter(
        models.Employee.id.in_(payload.employee_ids),
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False
    ).all()
    if not db_employees:
        raise HTTPException(status_code=404, detail="No matching employees found")

    fields_changed = []
    if payload.team is not None:
        fields_changed.append("team")
    if payload.department is not None:
        fields_changed.append("department")
    if payload.location is not None:
        fields_changed.append("location")
    if payload.manager is not None:
        fields_changed.append("manager")
    if payload.status is not None:
        fields_changed.append("status")
    if payload.hourly_rate_set is not None:
        fields_changed.append("hourly_rate (set)")
    if payload.hourly_rate_adjust_pct is not None:
        fields_changed.append("hourly_rate (adjust %)")

    for emp in db_employees:
        if payload.team is not None:
            emp.team = payload.team
        if payload.department is not None:
            emp.department = payload.department
        if payload.location is not None:
            emp.location = payload.location
        if payload.manager is not None:
            emp.manager = payload.manager
        if payload.status is not None:
            emp.status = payload.status
        if payload.hourly_rate_set is not None:
            emp.hourly_rate = payload.hourly_rate_set
        elif payload.hourly_rate_adjust_pct is not None:
            emp.hourly_rate = round(emp.hourly_rate * (1 + payload.hourly_rate_adjust_pct / 100.0), 2)

    db.commit()
    for emp in db_employees:
        db.refresh(emp)

    write_system_log(
        db, username=admin.username, action="Bulk Edit Employees",
        details=f"Applied {', '.join(fields_changed) or 'no changes'} to {len(db_employees)} employees: {', '.join(e.name for e in db_employees)}"
    )
    return db_employees


@app.delete("/api/employees/{employee_id}")
def delete_employee(employee_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    # Soft delete: the record (and its allocation history) moves to Trash
    # instead of being destroyed, so an accidental delete is recoverable.
    db_employee.is_deleted = True
    db_employee.deleted_at = datetime.datetime.utcnow()
    db.commit()
    write_system_log(db, username=admin.username, action="Delete Employee", details=f"Moved employee '{db_employee.name}' to Trash")
    return {"message": "Employee moved to Trash"}


@app.post("/api/employees/{employee_id}/restore", response_model=schemas.EmployeeResponse)
def restore_employee(employee_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    db_employee.is_deleted = False
    db_employee.deleted_at = None
    db.commit()
    db.refresh(db_employee)
    write_system_log(db, username=admin.username, action="Restore Employee", details=f"Restored employee '{db_employee.name}' from Trash")
    return db_employee

# ==========================================
# 3. TOPICS API (CRUD + Scenario-Scoped)
# ==========================================

@app.get("/api/topics")
def get_topics(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    topics = db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id,
        models.Topic.is_deleted == False
    ).all()
    
    # Include additional costs in topic structure for easy frontend consumption
    result = []
    for topic in topics:
        additional_costs_list = []
        for cost in topic.additional_costs:
            additional_costs_list.append({
                "id": cost.id,
                "cost_type": cost.cost_type,
                "category": cost.category,
                "amount": cost.amount,
                "notes": cost.notes
            })
            
        result.append({
            "id": topic.id,
            "name": topic.name,
            "category": topic.category,
            "area": topic.area,
            "description": topic.description,
            "objective": topic.objective,
            "deliverables": topic.deliverables,
            "justification": topic.justification,
            "status": topic.status,
            "comments": topic.comments,
            "notes": topic.notes,
            "recovery": topic.recovery,
            "additional_costs": additional_costs_list
        })
    return result

@app.post("/api/topics", response_model=schemas.TopicResponse)
def create_topic(topic: schemas.TopicCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    db_topic = models.Topic(
        scenario_id=active_scenario.id,
        name=topic.name,
        category=topic.category,
        area=topic.area,
        description=topic.description,
        objective=topic.objective,
        deliverables=topic.deliverables,
        justification=topic.justification,
        status=topic.status,
        comments=topic.comments,
        notes=topic.notes,
        recovery=topic.recovery
    )
    db.add(db_topic)
    db.commit()
    db.refresh(db_topic)
    return db_topic

@app.put("/api/topics/{topic_id}", response_model=schemas.TopicResponse)
def update_topic(topic_id: int, topic: schemas.TopicCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_topic = db.query(models.Topic).filter(models.Topic.id == topic_id).first()
    if not db_topic:
        raise HTTPException(status_code=404, detail="Topic not found")
        
    db_topic.name = topic.name
    db_topic.category = topic.category
    db_topic.area = topic.area
    db_topic.description = topic.description
    db_topic.objective = topic.objective
    db_topic.deliverables = topic.deliverables
    db_topic.justification = topic.justification
    db_topic.status = topic.status
    db_topic.comments = topic.comments
    db_topic.notes = topic.notes
    db_topic.recovery = topic.recovery
    
    db.commit()
    db.refresh(db_topic)
    return db_topic

@app.delete("/api/topics/{topic_id}")
def delete_topic(topic_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_topic = db.query(models.Topic).filter(models.Topic.id == topic_id).first()
    if not db_topic:
        raise HTTPException(status_code=404, detail="Topic not found")
    # Soft delete: moves to Trash (allocations/costs stay attached) instead of
    # being destroyed, so an accidental delete is recoverable.
    db_topic.is_deleted = True
    db_topic.deleted_at = datetime.datetime.utcnow()
    db.commit()
    write_system_log(db, username=admin.username, action="Delete Topic", details=f"Moved topic '{db_topic.name}' to Trash")
    return {"message": "Topic moved to Trash"}


@app.post("/api/topics/{topic_id}/restore", response_model=schemas.TopicResponse)
def restore_topic(topic_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_topic = db.query(models.Topic).filter(models.Topic.id == topic_id).first()
    if not db_topic:
        raise HTTPException(status_code=404, detail="Topic not found")
    db_topic.is_deleted = False
    db_topic.deleted_at = None
    db.commit()
    db.refresh(db_topic)
    write_system_log(db, username=admin.username, action="Restore Topic", details=f"Restored topic '{db_topic.name}' from Trash")
    return db_topic


@app.get("/api/trash", response_model=schemas.TrashResponse)
def get_trash(db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    deleted_employees = db.query(models.Employee).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == True
    ).order_by(models.Employee.deleted_at.desc()).all()
    deleted_topics = db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id,
        models.Topic.is_deleted == True
    ).order_by(models.Topic.deleted_at.desc()).all()
    return {"employees": deleted_employees, "topics": deleted_topics}


@app.delete("/api/employees/{employee_id}/permanent")
def permanently_delete_employee(employee_id: int, db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    db_employee = db.query(models.Employee).filter(models.Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not db_employee.is_deleted:
        raise HTTPException(status_code=400, detail="Employee must be in Trash before it can be permanently deleted")
    name = db_employee.name
    db.delete(db_employee)
    db.commit()
    write_system_log(db, username=master.username, action="Permanently Delete Employee", details=f"Permanently deleted employee '{name}' from Trash")
    return {"message": "Employee permanently deleted"}


@app.delete("/api/topics/{topic_id}/permanent")
def permanently_delete_topic(topic_id: int, db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    db_topic = db.query(models.Topic).filter(models.Topic.id == topic_id).first()
    if not db_topic:
        raise HTTPException(status_code=404, detail="Topic not found")
    if not db_topic.is_deleted:
        raise HTTPException(status_code=400, detail="Topic must be in Trash before it can be permanently deleted")
    name = db_topic.name
    db.delete(db_topic)
    db.commit()
    write_system_log(db, username=master.username, action="Permanently Delete Topic", details=f"Permanently deleted topic '{name}' from Trash")
    return {"message": "Topic permanently deleted"}


@app.delete("/api/trash")
def empty_trash(db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    active_scenario = get_active_scenario_db(db)
    deleted_employees = db.query(models.Employee).filter(
        models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == True
    ).all()
    deleted_topics = db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == True
    ).all()
    emp_count = len(deleted_employees)
    top_count = len(deleted_topics)
    for e in deleted_employees:
        db.delete(e)
    for t in deleted_topics:
        db.delete(t)
    db.commit()
    write_system_log(db, username=master.username, action="Empty Trash", details=f"Permanently deleted {emp_count} employee(s) and {top_count} topic(s) from Trash")
    return {"message": f"Trash emptied: {emp_count} employee(s) and {top_count} topic(s) permanently deleted"}

# ==========================================
# 3.1 TOPICS ADDITIONAL COSTS API
# ==========================================

@app.post("/api/topics/{topic_id}/costs", response_model=schemas.AdditionalCostResponse)
def add_topic_cost(topic_id: int, cost: schemas.AdditionalCostCreate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    topic = db.query(models.Topic).filter(models.Topic.id == topic_id).first()
    if not topic:
        raise HTTPException(status_code=404, detail="Topic not found")
        
    db_cost = models.AdditionalCost(
        topic_id=topic_id,
        cost_type=cost.cost_type,
        category=cost.category,
        amount=cost.amount,
        notes=cost.notes
    )
    db.add(db_cost)
    db.commit()
    db.refresh(db_cost)
    return db_cost

@app.delete("/api/topics/costs/{cost_id}")
def delete_topic_cost(cost_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_cost = db.query(models.AdditionalCost).filter(models.AdditionalCost.id == cost_id).first()
    if not db_cost:
        raise HTTPException(status_code=404, detail="Additional cost not found")
    db.delete(db_cost)
    db.commit()
    return {"message": "Additional cost deleted successfully"}

# ==========================================
# 4. ALLOCATIONS API (Grid-scoped)
# ==========================================

@app.get("/api/allocations", response_model=List[schemas.AllocationResponse])
def get_allocations(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    # Join with employees/topics to filter by scenario and hide Trash on both sides
    return db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()

@app.post("/api/allocations", response_model=schemas.AllocationResponse)
def save_allocation(alloc: schemas.AllocationUpdate, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    db_alloc = db.query(models.Allocation).filter(
        models.Allocation.employee_id == alloc.employee_id,
        models.Allocation.topic_id == alloc.topic_id
    ).first()
    
    if db_alloc:
        db_alloc.percentage = alloc.percentage
        if alloc.comment is not None:
            db_alloc.comment = alloc.comment
    else:
        db_alloc = models.Allocation(
            employee_id=alloc.employee_id,
            topic_id=alloc.topic_id,
            percentage=alloc.percentage,
            comment=alloc.comment
        )
        db.add(db_alloc)
        
    db.commit()
    db.refresh(db_alloc)
    return db_alloc

# ==========================================
# 5. DASHBOARDS & REPORTS AGGREGATION ENGINE
# ==========================================

@app.get("/api/reports/dashboard")
def get_dashboard_reports(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)

    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()

    # Pre-map allocations for performance
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()
    alloc_map = {} # (emp_id, topic_id) -> percentage
    emp_alloc_sums = {} # emp_id -> total percentage sum
    
    for a in allocations:
        alloc_map[(a.employee_id, a.topic_id)] = a.percentage
        emp_alloc_sums[a.employee_id] = emp_alloc_sums.get(a.employee_id, 0.0) + a.percentage

    total_headcount = len(employees)
    
    # 1. Calculate Employee Internal Costs
    total_internal_employee_cost = 0.0
    employee_costs_by_topic = {} # topic_id -> cost
    employee_costs_by_team = {} # team -> cost
    employee_costs_by_location = {} # location -> cost
    employee_costs_by_dept = {} # department -> cost
    
    overloaded_employees = []
    
    for emp in employees:
        utilization = emp_alloc_sums.get(emp.id, 0.0)
        if utilization > 100.0:
            overloaded_employees.append({
                "id": emp.id,
                "name": emp.name,
                "team": emp.team,
                "location": emp.location,
                "utilization": utilization
            })
            
        for topic in topics:
            pct = alloc_map.get((emp.id, topic.id), 0.0)
            if pct > 0.0:
                cost = emp.available_hours * emp.hourly_rate * (pct / 100.0)
                total_internal_employee_cost += cost
                
                employee_costs_by_topic[topic.id] = employee_costs_by_topic.get(topic.id, 0.0) + cost
                employee_costs_by_team[emp.team] = employee_costs_by_team.get(emp.team, 0.0) + cost
                employee_costs_by_location[emp.location] = employee_costs_by_location.get(emp.location, 0.0) + cost
                employee_costs_by_dept[emp.department] = employee_costs_by_dept.get(emp.department, 0.0) + cost

    # 2. Additional Costs Aggregation
    total_additional_internal = 0.0
    total_external_cost = 0.0
    total_recovery = 0.0
    
    topic_additional_costs = {} # topic_id -> {"internal": val, "external": val}
    
    for topic in topics:
        total_recovery += topic.recovery
        topic_additional_costs[topic.id] = {"internal": 0.0, "external": 0.0}
        
        for cost in topic.additional_costs:
            if cost.cost_type == "internal":
                total_additional_internal += cost.amount
                topic_additional_costs[topic.id]["internal"] += cost.amount
            elif cost.cost_type == "external":
                total_external_cost += cost.amount
                topic_additional_costs[topic.id]["external"] += cost.amount
                
    total_annual_planning_cost = total_internal_employee_cost + total_additional_internal + total_external_cost - total_recovery

    # 3. Topic Summaries
    topic_summaries = []
    cost_by_category = {}
    
    for topic in topics:
        emp_cost = employee_costs_by_topic.get(topic.id, 0.0)
        add_int = topic_additional_costs.get(topic.id, {}).get("internal", 0.0)
        ext_cost = topic_additional_costs.get(topic.id, {}).get("external", 0.0)
        total_topic_cost = emp_cost + add_int + ext_cost - topic.recovery
        
        # Get employees involved
        involved_staff = []
        for emp in employees:
            pct = alloc_map.get((emp.id, topic.id), 0.0)
            if pct > 0.0:
                involved_staff.append({
                    "employee_name": emp.name,
                    "team": emp.team,
                    "location": emp.location,
                    "percentage": pct,
                    "cost": emp.available_hours * emp.hourly_rate * (pct / 100.0)
                })
                
        topic_summaries.append({
            "id": topic.id,
            "name": topic.name,
            "category": topic.category,
            "area": topic.area,
            "status": topic.status,
            "employee_cost": emp_cost,
            "additional_internal_cost": add_int,
            "external_cost": ext_cost,
            "recovery": topic.recovery,
            "total_cost": total_topic_cost,
            "staff": involved_staff,
            "justification": topic.justification,
            "objective": topic.objective,
            "deliverables": topic.deliverables
        })
        
        cost_by_category[topic.category] = cost_by_category.get(topic.category, 0.0) + total_topic_cost

    # Sort topics by highest cost
    highest_cost_topics = sorted(topic_summaries, key=lambda x: x["total_cost"], reverse=True)[:5]
    
    # 4. Team summaries
    team_summaries = []
    unique_teams = set(emp.team for emp in employees)
    for team in unique_teams:
        team_members = [e for e in employees if e.team == team]
        team_emp_ids = [e.id for e in team_members]
        
        # Calculate team total cost and average utilization
        team_cost = employee_costs_by_team.get(team, 0.0)
        utils = [emp_alloc_sums.get(eid, 0.0) for eid in team_emp_ids]
        avg_util = sum(utils) / len(utils) if utils else 0.0
        over_limit = len([u for u in utils if u > 100.0])
        
        # Split topics where team contributes
        contributing_topics = []
        for t in topics:
            team_pct = sum(alloc_map.get((eid, t.id), 0.0) for eid in team_emp_ids)
            if team_pct > 0.0:
                contributing_topics.append({
                    "topic_name": t.name,
                    "total_percentage": team_pct,
                    "generated_cost": sum(emp.available_hours * emp.hourly_rate * (alloc_map.get((emp.id, t.id), 0.0) / 100.0) for emp in team_members)
                })
                
        team_summaries.append({
            "team_name": team,
            "member_count": len(team_members),
            "total_cost": team_cost,
            "average_utilization": avg_util,
            "overloaded_count": over_limit,
            "topics": contributing_topics
        })

    return {
        "scenario_name": active_scenario.name,
        "total_headcount": total_headcount,
        "total_internal_employee_cost": total_internal_employee_cost,
        "total_additional_internal_cost": total_additional_internal,
        "total_external_cost": total_external_cost,
        "total_recovery_cost": total_recovery,
        "total_annual_planning_cost": total_annual_planning_cost,
        "cost_by_location": employee_costs_by_location,
        "cost_by_team": employee_costs_by_team,
        "cost_by_department": employee_costs_by_dept,
        "cost_by_category": cost_by_category,
        "highest_cost_topics": highest_cost_topics,
        "overloaded_employees": overloaded_employees,
        "topic_summaries": topic_summaries,
        "team_summaries": team_summaries
    }

# ==========================================
# 6. SMART EXCEL / CSV IMPORTER
# ==========================================

COST_ROW_KEYWORDS = ["CAD", "Sampling", "Equipment", "PTF", "Testing", "Tooling", "Prototypes",
                      "Supplier", "Services", "Recovery", "Engineering", "Internal", "External", "Other"]

def classify_cost_category(cost_category: str) -> str:
    cat_lower = cost_category.lower()
    if "recovery" in cat_lower:
        return "recovery"
    if cat_lower.startswith("internal") or cat_lower in ("cad", "engineering support", "sampling", "ptf"):
        return "internal"
    return "external"

def parse_uploaded_file_to_rows(contents: bytes, filename: str, content_type: Optional[str]):
    """Parse raw CSV/Excel bytes into a list of string rows. Returns (rows, file_type)."""
    filename = (filename or "").lower()
    rows = []

    is_excel = (filename.endswith(".xlsx") or filename.endswith(".xls")
                or content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"])

    if is_excel:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                # Filter out completely empty rows
                if not any(v is not None for v in row):
                    continue
                row_vals = []
                for val in row:
                    if val is None:
                        row_vals.append("")
                    else:
                        row_vals.append(str(val))
                rows.append(row_vals)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
        file_type = "excel"
    else:
        try:
            decoded = contents.decode("utf-8")
            csv_file = io.StringIO(decoded)
            reader = csv.reader(csv_file)
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")
        file_type = "csv"

    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    return rows, file_type


CANONICAL_FIELD_ALIASES = {
    "employee": ["Employee", "Employee Name", "Name"],
    "team": ["Team"],
    "location": ["Location"],
    "hours": ["Hours / Year", "Available hours/year (Region 100%)", "Hours/Year"],
    "rate": ["Hourly Rate", "Hourly rate"],
    "department": ["Department", "Dept", "Area"],
    "manager": ["Manager"],
    "notes": ["Notes"],
    "status": ["Status"],
}
CANONICAL_FIELD_LABELS = {
    "employee": "Employee", "team": "Team", "location": "Location",
    "hours": "Hours / Year", "rate": "Hourly Rate", "department": "Department",
    "manager": "Manager", "notes": "Notes", "status": "Status",
}


def resolve_header_indices(headers: list, column_mapping: Optional[dict] = None) -> dict:
    """Maps canonical meta fields (employee, team, rate, ...) to their column
    index in `headers`, using known exact-name aliases first, then any
    admin-confirmed AI-suggested overrides in `column_mapping` (header text ->
    canonical field key)."""
    def find_idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    idxs = {field: find_idx(*aliases) for field, aliases in CANONICAL_FIELD_ALIASES.items()}

    for header, field in (column_mapping or {}).items():
        if header in headers and field in idxs:
            idxs[field] = headers.index(header)

    return idxs


def suggest_header_mappings(headers: list, idxs: dict, existing_topic_names: list) -> list:
    """For headers that didn't exactly match a known meta field and don't match
    an existing topic name either, fuzzy-suggest the closest meta field. This
    surfaces likely-mistyped columns (e.g. 'Hrly Rate') for admin confirmation
    instead of silently importing them as a brand-new topic full of numbers."""
    matched_idxs = {i for i in idxs.values() if i != -1}
    existing_topic_names_lower = {n.lower() for n in existing_topic_names}
    unmatched_fields = [field for field, i in idxs.items() if i == -1]

    suggestions = []
    for i, h in enumerate(headers):
        if i in matched_idxs or not h:
            continue
        h_lower = h.lower()
        if h_lower in existing_topic_names_lower:
            continue
        if "total" in h_lower or "utilization" in h_lower:
            continue

        best_field, best_score = None, 0.0
        for field in unmatched_fields:
            label_lower = CANONICAL_FIELD_LABELS[field].lower()
            score = difflib.SequenceMatcher(None, h_lower, label_lower).ratio()
            # Abbreviations like "Loc" for "Location" are a strong signal even
            # when the character-level ratio alone falls short of the threshold.
            if len(h_lower) >= 2 and (h_lower in label_lower or label_lower in h_lower):
                score += 0.15
            if score > best_score:
                best_field, best_score = field, score

        if best_field and best_score >= 0.55:
            suggestions.append({
                "header": h,
                "suggested_field": best_field,
                "suggested_label": CANONICAL_FIELD_LABELS[best_field],
                "confidence": round(best_score, 2)
            })

    return suggestions


def apply_rows_to_scenario(rows, active_scenario, db: Session, column_mapping: Optional[dict] = None):
    """Core matrix-sheet importer: parses header-driven rows into Employees/
    Topics/Allocations/AdditionalCosts for the given scenario. Shared by the
    direct file upload endpoint and the "apply from upload history" endpoint,
    so both behave identically. Returns a dict of import counts."""

    # Standard header checking
    headers = [h.strip() for h in rows[0]]

    # Only "Employee" is a hard requirement. Every other known column is optional:
    # if a sheet is missing one, existing records simply keep their current value
    # for that field instead of the whole import being rejected.
    idxs = resolve_header_indices(headers, column_mapping)
    emp_idx = idxs["employee"]
    if emp_idx == -1:
        raise HTTPException(status_code=400, detail="Required column 'Employee' not found in CSV headers.")

    team_idx = idxs["team"]
    loc_idx = idxs["location"]
    hours_idx = idxs["hours"]
    rate_idx = idxs["rate"]
    dept_idx = idxs["department"]
    manager_idx = idxs["manager"]
    notes_idx = idxs["notes"]
    status_idx = idxs["status"]

    known_meta_idxs = {i for i in idxs.values() if i != -1}

    # Any remaining column is treated as a topic/project allocation column - this is
    # what lets brand-new columns show up in the platform automatically, and lets
    # columns be reordered or omitted freely between uploads.
    topic_cols = []
    for idx, h in enumerate(headers):
        if idx in known_meta_idxs:
            continue
        if "total" in h.lower() or "utilization" in h.lower() or h == "":
            continue
        topic_cols.append((idx, h))

    # Clear current allocations and comments for active scenario
    db_employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id).all()
    emp_ids = [e.id for e in db_employees]
    db.query(models.Allocation).filter(models.Allocation.employee_id.in_(emp_ids)).delete(synchronize_session=False)
    # Delete additional costs for topics of active scenario
    additional_cost_ids = [
        ac.id for ac in db.query(models.AdditionalCost)
        .join(models.Topic)
        .filter(models.Topic.scenario_id == active_scenario.id)
        .all()
    ]
    if additional_cost_ids:
        db.query(models.AdditionalCost).filter(models.AdditionalCost.id.in_(additional_cost_ids)).delete(synchronize_session=False)

    
    # Note: We keep employees and topics, but we'll update or create them.
    # To prevent duplicates, we map existing employees/topics. Soft-deleted
    # (Trash) records are excluded so a name matching a deleted employee
    # creates a fresh record instead of silently reviving the hidden one.
    existing_emp = {e.name: e for e in db.query(models.Employee).filter(
        models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()}
    existing_top = {t.name: t for t in db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()}
    
    added_emps = 0
    added_tops = 0
    added_allocs = 0
    added_costs = 0
    
    def cell(row, idx):
        if idx == -1 or idx >= len(row):
            return ""
        return row[idx].strip()

    # We will process each row
    for r_idx in range(1, len(rows)):
        row = rows[r_idx]
        if not row or len(row) <= emp_idx:
            continue

        first_cell = cell(row, emp_idx)
        row_team = cell(row, team_idx)
        row_location = cell(row, loc_idx)

        # Bottom rows for additional costs / recovery reuse the Employee column to
        # hold the cost category name (e.g. "CAD", "Tooling", "Recovery") and leave
        # Team/Location blank - that combination is what tells them apart from a
        # genuine employee row.
        is_cost_row = (not row_team and not row_location and first_cell
                       and any(kw in first_cell for kw in COST_ROW_KEYWORDS))

        if is_cost_row:
            cost_category = first_cell
            cost_type = classify_cost_category(cost_category)

            # Loop through topic columns
            for col_idx, topic_name in topic_cols:
                if col_idx < len(row) and row[col_idx].strip():
                    val_str = row[col_idx].strip().replace("$", "").replace(",", "").replace("%", "")
                    try:
                        amt = float(val_str)
                        if amt != 0.0:
                            # Find topic
                            topic = existing_top.get(topic_name)
                            if topic:
                                if cost_type == "recovery":
                                    # Recovery is stored positive in DB but subtracted in final cost
                                    topic.recovery = abs(amt)
                                else:
                                    db_cost = models.AdditionalCost(
                                        topic_id=topic.id,
                                        cost_type=cost_type,
                                        category=cost_category,
                                        amount=amt,
                                        notes="Imported from CSV"
                                    )
                                    db.add(db_cost)
                                    added_costs += 1
                    except ValueError:
                        pass
            continue

        if not first_cell:
            # Blank separator row
            continue

        # Parse standard employee row. Every field below is optional: a blank cell,
        # or a column that doesn't exist in this particular sheet, leaves the
        # existing value untouched on update, and falls back to a sane default
        # when creating a brand-new employee.
        emp_name = first_cell
        team = row_team
        location = row_location
        dept_raw = cell(row, dept_idx)
        manager_raw = cell(row, manager_idx)
        notes_raw = cell(row, notes_idx)
        status_raw = cell(row, status_idx)

        available_hours = None
        hours_raw = cell(row, hours_idx)
        if hours_raw:
            try:
                available_hours = float(hours_raw.replace(",", "").strip())
            except ValueError:
                available_hours = None

        hourly_rate = None
        rate_raw = cell(row, rate_idx)
        if rate_raw:
            try:
                hourly_rate = float(rate_raw.replace("$", "").replace(",", "").strip())
            except ValueError:
                hourly_rate = None

        # Create or update Employee
        employee = existing_emp.get(emp_name)
        if not employee:
            # Set default department based on Team or Department field
            dept = dept_raw if dept_raw else ("CAE" if "cae" in team.lower() else "Test" if "test" in team.lower() else "Management")
            employee = models.Employee(
                scenario_id=active_scenario.id,
                name=emp_name,
                team=team or "Unassigned",
                department=dept,
                location=location or "Unassigned",
                available_hours=available_hours if available_hours is not None else 1800.0,
                hourly_rate=hourly_rate if hourly_rate is not None else 50.0,
                status=status_raw or "Active",
                manager=manager_raw or None,
                notes=notes_raw or None
            )
            db.add(employee)
            db.commit()
            db.refresh(employee)
            existing_emp[emp_name] = employee
            added_emps += 1
        else:
            if team:
                employee.team = team
            if location:
                employee.location = location
            if available_hours is not None:
                employee.available_hours = available_hours
            if hourly_rate is not None:
                employee.hourly_rate = hourly_rate
            if dept_raw:
                employee.department = dept_raw
            if manager_raw:
                employee.manager = manager_raw
            if notes_raw:
                employee.notes = notes_raw
            if status_raw:
                employee.status = status_raw
            db.commit()

        # Loop through Topic Columns for allocations
        for col_idx, topic_name in topic_cols:
            if col_idx < len(row):
                pct_str = row[col_idx].strip().replace("%", "")
                if pct_str:
                    try:
                        pct = float(pct_str)
                        # Clean decimal percentages if input is, say, 0.2 instead of 20
                        if 0.0 < pct <= 1.0:
                            pct = pct * 100.0
                            
                        if pct > 0.0:
                            # Create Topic if missing
                            topic = existing_top.get(topic_name)
                            if not topic:
                                cat = "Internal Efforts" if "non bookable" in topic_name.lower() or "consulting" in topic_name.lower() else "Internal D-Projects" if "ai" in topic_name.lower() else "Customer Requests"
                                topic = models.Topic(
                                    scenario_id=active_scenario.id,
                                    name=topic_name,
                                    category=cat,
                                    area="CAE" if "cae" in topic_name.lower() else "Test" if "test" in topic_name.lower() else "General",
                                    status="Active"
                                )
                                db.add(topic)
                                db.commit()
                                db.refresh(topic)
                                existing_top[topic_name] = topic
                                added_tops += 1
                                
                            # Save allocation
                            db_alloc = models.Allocation(
                                employee_id=employee.id,
                                topic_id=topic.id,
                                percentage=pct,
                                comment="Imported from CSV"
                            )
                            db.add(db_alloc)
                            added_allocs += 1
                    except ValueError:
                        pass
                        
    db.commit()
    return {
        "imported_employees": added_emps,
        "imported_topics": added_tops,
        "imported_allocations": added_allocs,
        "imported_additional_costs": added_costs
    }


@app.post("/api/import/preview")
async def preview_import_columns(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    """AI-assisted column mapping: parses just the headers of an uploaded sheet
    and, for any column that doesn't exactly match a known field (Employee,
    Team, Hourly Rate, ...) or an existing topic name, fuzzy-suggests the
    closest known field. The frontend shows these as a confirmation step
    before the real import, so a mistyped column like 'Hrly Rate' doesn't
    silently become a bogus new topic."""
    active_scenario = get_active_scenario_db(db)
    contents = await file.read()
    rows, file_type = parse_uploaded_file_to_rows(contents, file.filename, file.content_type)
    headers = [h.strip() for h in rows[0]]
    idxs = resolve_header_indices(headers)

    existing_topics = db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False
    ).all()
    suggestions = suggest_header_mappings(headers, idxs, [t.name for t in existing_topics])

    return {"headers": headers, "suggested_mappings": suggestions}


@app.post("/api/import/csv")
async def import_csv_data(file: UploadFile = File(...), column_mapping: Optional[str] = Form(None), db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)

    contents = await file.read()
    rows, file_type = parse_uploaded_file_to_rows(contents, file.filename, file.content_type)

    mapping_dict = None
    if column_mapping:
        try:
            mapping_dict = json.loads(column_mapping)
        except (ValueError, TypeError):
            mapping_dict = None

    counts = apply_rows_to_scenario(rows, active_scenario, db, column_mapping=mapping_dict)

    # Keep a copy of the raw file on disk so it shows up in "Upload History"
    # and can be re-applied later without the user needing the original file.
    stored_filename = f"{uuid.uuid4().hex}_{os.path.basename(file.filename or 'upload')}"
    with open(os.path.join(UPLOAD_STORAGE_DIR, stored_filename), "wb") as f:
        f.write(contents)

    upload_record = models.UploadHistory(
        scenario_id=active_scenario.id,
        original_filename=file.filename or "upload",
        stored_filename=stored_filename,
        file_type=file_type,
        size_bytes=len(contents),
        uploaded_by=admin.username,
        imported_employees=counts["imported_employees"],
        imported_topics=counts["imported_topics"],
        imported_allocations=counts["imported_allocations"],
        imported_additional_costs=counts["imported_additional_costs"]
    )
    db.add(upload_record)
    db.commit()

    write_system_log(
        db,
        username=admin.username,
        action="Import CSV",
        details=f"Successfully imported '{file.filename}'. Loaded {counts['imported_employees']} employees, "
                f"{counts['imported_topics']} topics, {counts['imported_allocations']} allocations, {counts['imported_additional_costs']} costs."
    )
    return {
        "status": "success",
        "message": f"Successfully parsed and loaded planning sheet into '{active_scenario.name}'",
        **counts
    }


@app.get("/api/uploads/history", response_model=List[schemas.UploadHistoryResponse])
def get_upload_history(db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    return (
        db.query(models.UploadHistory)
        .filter(models.UploadHistory.scenario_id == active_scenario.id)
        .order_by(models.UploadHistory.uploaded_at.desc())
        .all()
    )


@app.delete("/api/uploads/history/{upload_id}")
def delete_upload_history(upload_id: int, db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    record = db.query(models.UploadHistory).filter(models.UploadHistory.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Upload history record not found")

    stored_path = os.path.join(UPLOAD_STORAGE_DIR, record.stored_filename)
    if os.path.exists(stored_path):
        os.remove(stored_path)

    filename = record.original_filename
    db.delete(record)
    db.commit()
    write_system_log(db, username=master.username, action="Delete Upload History", details=f"Permanently deleted upload history record for '{filename}'")
    return {"message": f"Deleted upload history record for '{filename}'"}


@app.post("/api/uploads/history/{upload_id}/apply")
def apply_upload_from_history(upload_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    record = db.query(models.UploadHistory).filter(models.UploadHistory.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Upload history record not found")

    stored_path = os.path.join(UPLOAD_STORAGE_DIR, record.stored_filename)
    if not os.path.exists(stored_path):
        raise HTTPException(status_code=404, detail="The stored file for this upload is no longer available")

    active_scenario = get_active_scenario_db(db)

    with open(stored_path, "rb") as f:
        contents = f.read()

    content_type = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                     if record.file_type == "excel" else "text/csv")
    rows, _ = parse_uploaded_file_to_rows(contents, record.original_filename, content_type)
    counts = apply_rows_to_scenario(rows, active_scenario, db)

    write_system_log(
        db,
        username=admin.username,
        action="Import CSV",
        details=f"Re-applied historical upload '{record.original_filename}' (originally uploaded {record.uploaded_at.strftime('%Y-%m-%d %H:%M')} UTC "
                f"by {record.uploaded_by}) onto '{active_scenario.name}'. Loaded {counts['imported_employees']} employees, "
                f"{counts['imported_topics']} topics, {counts['imported_allocations']} allocations, {counts['imported_additional_costs']} costs."
    )
    return {
        "status": "success",
        "message": f"Successfully re-applied '{record.original_filename}' onto '{active_scenario.name}'",
        **counts
    }


@app.get("/api/export/excel")
def export_excel_data(
    location: Optional[str] = None,
    team: Optional[str] = None,
    department: Optional[str] = None,
    category: Optional[str] = None,
    minRate: Optional[float] = None,
    maxRate: Optional[float] = None,
    manager: Optional[str] = None,
    status: Optional[str] = None,
    topicId: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    active_scenario = get_active_scenario_db(db)

    # Load data
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()
    
    alloc_map = {}
    for a in allocations:
        alloc_map[(a.employee_id, a.topic_id)] = a.percentage

    # Apply filters
    filtered_employees = []
    for emp in employees:
        if location and emp.location != location: continue
        if team and emp.team != team: continue
        if department and emp.department != department: continue
        if manager and emp.manager != manager: continue
        if status and emp.status != status: continue
        if minRate is not None and emp.hourly_rate < minRate: continue
        if maxRate is not None and emp.hourly_rate > maxRate: continue
        if topicId is not None:
            pct = alloc_map.get((emp.id, topicId), 0.0)
            if pct <= 0.0: continue
        filtered_employees.append(emp)
        
    filtered_topics = []
    for t in topics:
        if category and t.category != category: continue
        filtered_topics.append(t)
        
    # Build a styled workbook matching the app's own brand look (dark blue
    # header band, currency/percentage number formats, zebra-striped rows,
    # a distinct fill for the cost/recovery summary rows) rather than a bare
    # openpyxl default. The data reflects whatever filters are currently
    # active in the UI, since filtered_employees/filtered_topics above are
    # already narrowed by the same query params the matrix uses.
    BRAND_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    COST_ROW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    RECOVERY_ROW_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    THIN_BORDER = Border(*(Side(style="thin", color="CBD5E1") for _ in range(4)))
    TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
    SUBTITLE_FONT = Font(italic=True, size=10, color="E0E7FF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Matrix"

    num_cols = 5 + len(filtered_topics)
    last_col_letter = get_column_letter(num_cols)

    # Row 1: title banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "Textron Digital Engineering - Resource Allocation Matrix"
    title_cell.font = TITLE_FONT
    title_cell.fill = BRAND_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: subtitle - scenario name, export time, and active filters so the
    # sheet is self-documenting about which slice of the data it represents.
    active_filters = []
    if location: active_filters.append(f"Location={location}")
    if team: active_filters.append(f"Team={team}")
    if department: active_filters.append(f"Department={department}")
    if category: active_filters.append(f"Topic Category={category}")
    if manager: active_filters.append(f"Manager={manager}")
    if status: active_filters.append(f"Status={status}")
    if topicId is not None: active_filters.append(f"Active Project Allocation={topicId}")
    if minRate is not None: active_filters.append(f"Min Rate={minRate}")
    if maxRate is not None: active_filters.append(f"Max Rate={maxRate}")
    filters_text = f" | Filters: {', '.join(active_filters)}" if active_filters else " | No filters applied"

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"{active_scenario.name} - Exported {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}{filters_text}"
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.fill = BRAND_FILL
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: blank spacer
    header_row_idx = 4

    # Row 4: column headers
    headers = ["Employee", "Team", "Location", "Hours/Year", "Hourly Rate"]
    for t in filtered_topics:
        headers.append(t.name)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row_idx].height = 28

    # Employee rows
    row_idx = header_row_idx
    for i, emp in enumerate(filtered_employees):
        row_idx += 1
        values = [emp.name, emp.team, emp.location, emp.available_hours, emp.hourly_rate]
        for t in filtered_topics:
            values.append(alloc_map.get((emp.id, t.id), 0.0))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx == 4:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx > 5:
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center")

    # Additional costs
    category_rows = {}
    for t in filtered_topics:
        for ac in t.additional_costs:
            if ac.category not in category_rows:
                category_rows[ac.category] = {}
            category_rows[ac.category][t.id] = category_rows[ac.category].get(t.id, 0.0) + ac.amount

    has_recovery = any(t.recovery and t.recovery != 0 for t in filtered_topics)

    if category_rows or has_recovery:
        row_idx += 1  # blank separator row

    for cat_name, topic_costs in category_rows.items():
        row_idx += 1
        values = [cat_name, "", "", "", ""]
        for t in filtered_topics:
            values.append(topic_costs.get(t.id, ""))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = COST_ROW_FILL
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx > 5 and val != "":
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    if has_recovery:
        row_idx += 1
        values = ["Recovery", "", "", "", ""]
        for t in filtered_topics:
            values.append(t.recovery if t.recovery else "")
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = RECOVERY_ROW_FILL
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx > 5 and val != "":
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Column widths sized to content, and freeze the header row + Employee
    # column so both stay visible while scrolling a large matrix.
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    for col_idx, t in enumerate(filtered_topics, start=6):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(32, len(t.name) // 1.3))
    ws.freeze_panes = f"B{header_row_idx + 1}"

    # Save workbook to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers_resp = {
        'Content-Disposition': f'attachment; filename="Allocation_Matrix_{active_scenario.name.replace(" ", "_")}.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)

# ==========================================
# 7. CONFIDENTIAL LOCAL AI ASSISTANT
# ==========================================

@app.get("/api/reports/ai-predictions")
def get_ai_predictions(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()
    
    # Calculate utilization
    emp_alloc_sums = {}
    for a in allocations:
        emp_alloc_sums[a.employee_id] = emp_alloc_sums.get(a.employee_id, 0.0) + a.percentage
        
    overloaded = [e for e in employees if emp_alloc_sums.get(e.id, 0.0) > 100.0]
    
    predictions = {
        "bottlenecks": [],
        "cost_optimizations": [],
        "reallocations": []
    }
    
    # 1. Bottlenecks
    if overloaded:
        for emp in overloaded:
            total_pct = emp_alloc_sums[emp.id]
            predictions["bottlenecks"].append({
                "type": "Resource Overload",
                "severity": "High",
                "description": f"Employee **{emp.name}** is overloaded at **{total_pct:.1f}%** utilization. This predicts burn-out or project milestone delivery delays."
            })
    else:
        predictions["bottlenecks"].append({
            "type": "Resource Staffing",
            "severity": "Low",
            "description": "All employee allocations are within safe bounds (<= 100%). Overall staff utilization risk is low."
        })
        
    # Find underallocated critical projects
    alloc_topic_sums = {}
    for a in allocations:
        alloc_topic_sums[a.topic_id] = alloc_topic_sums.get(a.topic_id, 0.0) + a.percentage
        
    for t in topics:
        tot = alloc_topic_sums.get(t.id, 0.0)
        if tot > 0.0 and tot < 50.0:
            predictions["bottlenecks"].append({
                "type": "Underallocation Risk",
                "severity": "Medium",
                "description": f"Topic **{t.name}** has only **{tot:.1f}%** total planned resource allocation. Deliverables may be delayed due to insufficient staff load."
            })

    # 2. Cost Optimizations
    for t in topics:
        add_ext = sum(ac.amount for ac in t.additional_costs if ac.cost_type == "external")
        # calculate internal staff cost
        emp_cost = 0.0
        for emp in employees:
            alloc_pct = sum(a.percentage for a in allocations if a.employee_id == emp.id and a.topic_id == t.id)
            emp_cost += emp.available_hours * emp.hourly_rate * (alloc_pct / 100.0)
            
        if add_ext > emp_cost and emp_cost > 0:
            savings = add_ext * 0.15
            predictions["cost_optimizations"].append({
                "category": "External Scope Optimization",
                "impact": f"Save ~${savings:,.2f}",
                "description": f"Topic **{t.name}** has high external vendor tooling/costs (${add_ext:,.2f}) compared to internal employee cost (${emp_cost:,.2f}). Suggest sampling internally to cut costs."
            })
            
    if not predictions["cost_optimizations"]:
        predictions["cost_optimizations"].append({
            "category": "Portfolio Savings",
            "impact": "Save ~$12,500.00",
            "description": "Review recovery rates on internal initiatives. Standardizing equipment sampling rates across Germany and Romania hubs is predicted to recover 15% budget."
        })

    # 3. Staff Reallocations
    if overloaded:
        for emp in overloaded:
            total_pct = emp_alloc_sums[emp.id]
            excess = total_pct - 100.0
            same_team_under = [e for e in employees if e.team == emp.team and emp_alloc_sums.get(e.id, 0.0) < 80.0]
            if same_team_under:
                helper = same_team_under[0]
                predictions["reallocations"].append({
                    "action": "Balance Staff Load",
                    "priority": "High",
                    "description": f"Reallocate **{excess:.1f}%** of load from overloaded **{emp.name}** to **{helper.name}** ({helper.team}) who has available bandwidth (currently at {emp_alloc_sums.get(helper.id, 0.0):.1f}%)."
                })
            else:
                predictions["reallocations"].append({
                    "action": "Hire External Vendor Support",
                    "priority": "Medium",
                    "description": f"Deploy external sampling contract to cover the excess **{excess:.1f}%** workload for **{emp.name}** to mitigate milestone delays."
                })
    else:
        predictions["reallocations"].append({
            "action": "Resource Optimization",
            "priority": "Low",
            "description": "No reallocations needed. Bandwidths are balanced across all planned employees."
        })
        
    predictions = enrich_predictions_with_llm(predictions, employees, topics, allocations, emp_alloc_sums, alloc_topic_sums)
    return predictions


def enrich_predictions_with_llm(predictions: dict, employees: list, topics: list, allocations: list, emp_alloc_sums: dict, alloc_topic_sums: dict) -> dict:
    """Feeds the same real, already-computed portfolio data (utilization,
    underallocated topics) into the local LLM for richer, less templated
    Admin Insights narrative. Pure enrichment: silently falls back to the
    existing heuristic-only predictions if no LLM is reachable or its output
    isn't valid JSON, so this is never a hard dependency for the endpoint."""
    top_util = sorted(
        [(e.name, e.team, emp_alloc_sums.get(e.id, 0.0)) for e in employees],
        key=lambda x: x[2], reverse=True
    )[:8]
    under_topics = sorted(
        [(t.name, t.category, alloc_topic_sums.get(t.id, 0.0)) for t in topics],
        key=lambda x: x[2]
    )[:8]

    data_summary = (
        "Employee utilization (name, team, total allocation %):\n" +
        "\n".join(f"- {n} ({team}): {pct:.1f}%" for n, team, pct in top_util) +
        "\n\nTopic total allocation (name, category, total allocation %):\n" +
        "\n".join(f"- {n} ({cat}): {pct:.1f}%" for n, cat, pct in under_topics)
    )

    prompt = (
        "You are a resource planning analyst for Textron. Based ONLY on the real portfolio data below, "
        "suggest up to 2 additional bottleneck risks, up to 2 additional cost optimizations, and up to 2 "
        "additional staff reallocation actions that are NOT obvious duplicates of simple overload/underload "
        "statements. Be specific and reference the real names/topics given.\n\n"
        f"Portfolio Data:\n{data_summary}\n\n"
        "Respond with STRICT JSON only, no prose, matching exactly this shape:\n"
        '{"bottlenecks": [{"type": "...", "severity": "High|Medium|Low", "description": "..."}], '
        '"cost_optimizations": [{"category": "...", "impact": "...", "description": "..."}], '
        '"reallocations": [{"action": "...", "priority": "High|Medium|Low", "description": "..."}]}'
    )

    raw = query_local_ollama(prompt)
    if not raw:
        return predictions

    try:
        # Ollama sometimes wraps JSON in prose/code fences despite instructions - extract the outermost braces.
        start = raw.index("{")
        end = raw.rindex("}") + 1
        llm_data = json.loads(raw[start:end])
    except (ValueError, json.JSONDecodeError):
        return predictions

    for key, required_keys in (
        ("bottlenecks", {"type", "severity", "description"}),
        ("cost_optimizations", {"category", "impact", "description"}),
        ("reallocations", {"action", "priority", "description"}),
    ):
        items = llm_data.get(key)
        if not isinstance(items, list):
            continue
        for item in items[:2]:
            if isinstance(item, dict) and required_keys.issubset(item.keys()):
                predictions[key].append(item)

    return predictions


def fuzzy_match_ambiguous(query: str, choices: list, key_extractor, threshold: float = 0.25) -> tuple[Optional[any], list[any]]:
    if not query:
        return None, []
    query_clean = query.lower().strip()
    query_tokens = set(query_clean.split())
    if not query_tokens:
        return None, []
        
    scored_items = []
    
    for item in choices:
        val = key_extractor(item).lower()
        val_tokens = set(val.split())
        
        # Jaccard overlap
        intersection = query_tokens.intersection(val_tokens)
        union = query_tokens.union(val_tokens)
        jaccard = len(intersection) / len(union) if union else 0
        
        # Word containment
        containment = sum(1.0 for q_t in query_tokens if q_t in val) / len(query_tokens)
        
        score = jaccard * 0.6 + containment * 0.4
        
        # Direct substring containment boost
        if query_clean in val:
            score += 0.5
            
        # Exact match boost
        if query_clean == val:
            score += 1.0
            
        if score >= threshold:
            scored_items.append((score, item))
            
    if not scored_items:
        return None, []
        
    # Sort descending
    scored_items.sort(key=lambda x: x[0], reverse=True)
    
    # Exact match override
    if scored_items[0][0] >= 1.4:
        return scored_items[0][1], []
        
    # Only one candidate
    if len(scored_items) == 1:
        return scored_items[0][1], []
        
    # Large score difference (clear winner)
    if scored_items[0][0] - scored_items[1][0] > 0.35:
        return scored_items[0][1], []
        
    # Return all candidates within range of top score
    top_score = scored_items[0][0]
    candidates = [item for score, item in scored_items if top_score - score <= 0.3]
    
    if len(candidates) == 1:
        return candidates[0], []
        
    return None, candidates


def _apply_ai_allocation(db: Session, employee_id: int, topic_id: int, percentage: float):
    db_alloc = db.query(models.Allocation).filter(
        models.Allocation.employee_id == employee_id,
        models.Allocation.topic_id == topic_id
    ).first()
    if db_alloc:
        db_alloc.percentage = percentage
    else:
        db_alloc = models.Allocation(employee_id=employee_id, topic_id=topic_id, percentage=percentage)
        db.add(db_alloc)
    db.commit()


AI_ACTION_PREFIX = "[Local Heuristic Engine Fallback] "


def try_execute_ai_action(q_lower: str, employees: list, topics: list, alloc_map: dict, db: Session, current_user: models.User) -> Optional[dict]:
    """Recognizes allocation-change requests in chat (e.g. 'set X's allocation on
    Y to 30%' or 'move 10% of X's Y time to Z') and executes them directly,
    instead of only describing what the user should do manually."""
    emp_part = topic_part = pct_part = None

    set_match = re.search(r"set\s+(.+?)'s\s+allocation\s+(?:on|for)\s+(.+?)\s+to\s+(\d+(?:\.\d+)?)\s*%?", q_lower)
    if set_match:
        emp_part, topic_part, pct_part = set_match.groups()
    else:
        set_match2 = re.search(r"set\s+(.+?)'s\s+(.+?)\s+allocation\s+to\s+(\d+(?:\.\d+)?)\s*%?", q_lower)
        if set_match2:
            emp_part, topic_part, pct_part = set_match2.groups()

    move_match = None
    if emp_part is None:
        move_match = re.search(r"move\s+(\d+(?:\.\d+)?)\s*%?\s+of\s+(.+?)'s\s+(.+?)\s+(?:time\s+)?to\s+(.+)", q_lower)

    if emp_part is None and not move_match:
        return None

    if current_user.role not in ("admin", "master_admin"):
        return {"answer": AI_ACTION_PREFIX + "I recognized an allocation-change request, but only Admins can execute changes via chat. Please ask an Admin, or apply it manually in the Allocation Matrix."}

    if move_match:
        pct_str, emp_part, topic_from_part, topic_to_part = move_match.groups()
        pct_delta = float(pct_str)

        emp_match, _ = fuzzy_match_ambiguous(emp_part.strip(), employees, lambda e: e.name)
        if not emp_match:
            return {"answer": AI_ACTION_PREFIX + f"I couldn't uniquely identify the employee '{emp_part.strip()}' to move allocation for."}
        topic_from_match, _ = fuzzy_match_ambiguous(topic_from_part.strip(), topics, lambda t: t.name)
        topic_to_match, _ = fuzzy_match_ambiguous(topic_to_part.strip(), topics, lambda t: t.name)
        if not topic_from_match or not topic_to_match:
            return {"answer": AI_ACTION_PREFIX + f"I couldn't identify both topics ('{topic_from_part.strip()}' -> '{topic_to_part.strip()}') to move allocation between."}

        current_from = alloc_map.get((emp_match.id, topic_from_match.id), 0.0)
        if pct_delta > current_from:
            return {"answer": AI_ACTION_PREFIX + f"{emp_match.name} is only allocated {current_from:.1f}% to {topic_from_match.name}, so I can't move {pct_delta:.1f}%."}

        new_from = round(current_from - pct_delta, 2)
        current_to = alloc_map.get((emp_match.id, topic_to_match.id), 0.0)
        new_to = round(current_to + pct_delta, 2)

        _apply_ai_allocation(db, emp_match.id, topic_from_match.id, new_from)
        _apply_ai_allocation(db, emp_match.id, topic_to_match.id, new_to)
        write_system_log(
            db, username=current_user.username, action="AI Chat Allocation Change",
            details=f"Moved {pct_delta:.1f}% of {emp_match.name}'s allocation from '{topic_from_match.name}' to '{topic_to_match.name}' via AI chat"
        )
        return {
            "answer": AI_ACTION_PREFIX + f"Done. Moved **{pct_delta:.1f}%** of **{emp_match.name}**'s time from **{topic_from_match.name}** ({current_from:.1f}% → {new_from:.1f}%) to **{topic_to_match.name}** ({current_to:.1f}% → {new_to:.1f}%). The Allocation Matrix has been updated.",
            "action_executed": True
        }

    pct = float(pct_part)
    if pct < 0 or pct > 100:
        return {"answer": AI_ACTION_PREFIX + "Allocation percentage must be between 0 and 100."}

    emp_match, _ = fuzzy_match_ambiguous(emp_part.strip(), employees, lambda e: e.name)
    if not emp_match:
        return {"answer": AI_ACTION_PREFIX + f"I couldn't uniquely identify the employee '{emp_part.strip()}'."}
    topic_match, _ = fuzzy_match_ambiguous(topic_part.strip(), topics, lambda t: t.name)
    if not topic_match:
        return {"answer": AI_ACTION_PREFIX + f"I couldn't uniquely identify the topic/project '{topic_part.strip()}'."}

    current_pct = alloc_map.get((emp_match.id, topic_match.id), 0.0)
    _apply_ai_allocation(db, emp_match.id, topic_match.id, pct)
    write_system_log(
        db, username=current_user.username, action="AI Chat Allocation Change",
        details=f"Set {emp_match.name}'s allocation on '{topic_match.name}' to {pct:.1f}% via AI chat (was {current_pct:.1f}%)"
    )
    return {
        "answer": AI_ACTION_PREFIX + f"Done. Set **{emp_match.name}**'s allocation on **{topic_match.name}** to **{pct:.1f}%** (was {current_pct:.1f}%). The Allocation Matrix has been updated.",
        "action_executed": True
    }


@app.post("/api/ai/query")
def local_ai_query(payload: dict, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    query = payload.get("query", "").strip()
    history = payload.get("history", [])
    if not query:
        raise HTTPException(status_code=400, detail="Query is empty")
        
    # Strict Guardrails & Rules Checks (Scope validation)
    out_of_scope_keywords = ["recipe", "capital of", "weather", "translate", "how to build a", "write a code", "python script", "javascript script", "history of", "who is the president", "poem", "joke"]
    planning_words = ["cost", "employee", "topic", "project", "allocat", "team", "recovery", "utiliz", "hour", "rate", "overload", "justif", "area", "cae", "test", "where", "who", "what", "budget"]
    
    q_lower = query.lower()
    is_out_of_scope = any(keyword in q_lower for keyword in out_of_scope_keywords)
    
    # Only enforce planning terms restriction for the initial query (history is empty)
    has_no_planning_terms = False
    if not history:
        has_no_planning_terms = not any(word in q_lower for word in planning_words)
    
    if is_out_of_scope or has_no_planning_terms:
        return {
            "answer": "I am a confidential resource planning assistant for Textron. "
                      "Under strict local guardrails, I am only authorized to answer planning, staffing, cost calculations, "
                      "and allocations questions based on the active scenario."
        }
        
    active_scenario = get_active_scenario_db(db)
    
    # 2. Fallback to smart pattern matching (Heuristic Local Engine)
    # We load database tables early to resolve both Heuristics and potential Follow-up ambiguities
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()

    alloc_map = {}
    emp_alloc_sums = {}
    for a in allocations:
        alloc_map[(a.employee_id, a.topic_id)] = a.percentage
        emp_alloc_sums[a.employee_id] = emp_alloc_sums.get(a.employee_id, 0.0) + a.percentage
        
    prefix = "[Local Heuristic Engine Fallback] "

    # Let the AI directly execute an allocation-change request (e.g. "set X's
    # allocation on Y to 30%" or "move 10% of X's Y time to Z") instead of
    # only describing what the user should do manually.
    action_result = try_execute_ai_action(q_lower, employees, topics, alloc_map, db, current_user)
    if action_result:
        return action_result

    # Check for reset/clear filters first
    q_clean = q_lower.replace("rate higher than", "rate >").replace("rate lower than", "rate <").replace("hourly rate >", "rate >").replace("hourly rate <", "rate <").replace("rate above", "rate >").replace("rate below", "rate <").replace("rate more than", "rate >").replace("rate less than", "rate <")
    
    if "clear filter" in q_clean or "reset filter" in q_clean or "show all" in q_clean:
        return {
            "answer": prefix + "I have reset all matrix filters to display the full dataset.",
            "filters": {
                "location": "",
                "team": "",
                "department": "",
                "category": "",
                "minRate": 0,
                "maxRate": 999999
            }
        }

    # Extract location, team, department, category, rate filters
    filter_loc = None
    unique_locs = list(set(e.location for e in employees))
    for loc in unique_locs:
        if loc.lower() in q_clean:
            filter_loc = loc
            break
            
    filter_team = None
    unique_teams = list(set(e.team for e in employees))
    for team in unique_teams:
        if team.lower() in q_clean:
            filter_team = team
            break
            
    filter_dept = None
    unique_depts = list(set(e.department for e in employees))
    for dept in unique_depts:
        if dept.lower() in q_clean:
            filter_dept = dept
            break
            
    filter_cat = None
    unique_cats = list(set(t.category for t in topics))
    for cat in unique_cats:
        if cat.lower() in q_clean:
            filter_cat = cat
            break
            
    filter_min_rate = None
    filter_max_rate = None
    
    # check rate > X or rate higher than X
    rate_gt = re.search(r"rate\s*>\s*(\d+)", q_clean)
    if rate_gt:
        filter_min_rate = float(rate_gt.group(1))
    else:
        # Check text: "higher than 70" or "above 70" or "more than 70" or "greater than 70" or "over 70"
        rate_gt_text = re.search(r"(?:higher than|above|greater than|more than|over)\s*(\d+)", q_clean)
        if rate_gt_text:
            filter_min_rate = float(rate_gt_text.group(1))
            
    rate_lt = re.search(r"rate\s*<\s*(\d+)", q_clean)
    if rate_lt:
        filter_max_rate = float(rate_lt.group(1))
    else:
        # Check text: "lower than 70" or "below 70" or "less than 70" or "under 70"
        rate_lt_text = re.search(r"(?:lower than|below|less than|under)\s*(\d+)", q_clean)
        if rate_lt_text:
            filter_max_rate = float(rate_lt_text.group(1))

    # Trigger filter application if any valid planning property is matched
    if filter_loc or filter_team or filter_dept or filter_cat or filter_min_rate is not None or filter_max_rate is not None:
        filters_to_apply = {}
        msg_parts = []
        
        # Calculate matching list to output in chat bubble
        matching_emps = []
        for e in employees:
            if filter_loc and e.location.lower() != filter_loc.lower(): continue
            if filter_team and e.team.lower() != filter_team.lower(): continue
            if filter_dept and e.department.lower() != filter_dept.lower(): continue
            if filter_min_rate is not None and e.hourly_rate < filter_min_rate: continue
            if filter_max_rate is not None and e.hourly_rate > filter_max_rate: continue
            matching_emps.append(e)
            
        if filter_loc:
            filters_to_apply["location"] = filter_loc
            msg_parts.append(f"Location: **{filter_loc}**")
        if filter_team:
            filters_to_apply["team"] = filter_team
            msg_parts.append(f"Team: **{filter_team}**")
        if filter_dept:
            filters_to_apply["department"] = filter_dept
            msg_parts.append(f"Department: **{filter_dept}**")
        if filter_cat:
            filters_to_apply["category"] = filter_cat
            msg_parts.append(f"Topic Category: **{filter_cat}**")
        if filter_min_rate is not None:
            filters_to_apply["minRate"] = filter_min_rate
            msg_parts.append(f"Hourly Rate: **> {filter_min_rate} USD**")
        if filter_max_rate is not None:
            filters_to_apply["maxRate"] = filter_max_rate
            msg_parts.append(f"Hourly Rate: **< {filter_max_rate} USD**")
            
        filters_list_str = ", ".join(msg_parts)
        
        if matching_emps:
            staff_details = []
            for e in matching_emps:
                staff_details.append(f"* **{e.name}** ({e.team}, {e.location}) - Hourly Rate: ${e.hourly_rate}/hr, Utilization: {emp_alloc_sums.get(e.id, 0.0):.1f}%")
            nl = "\n"
            answer_text = (
                f"{prefix}Applied filters to Allocation Matrix: {filters_list_str}.\n\n"
                f"Here are the **{len(matching_emps)}** matching employees currently visualised in the matrix:\n"
                f"{nl.join(staff_details)}\n\n"
                f"The Resource Allocation Matrix has been updated in the background. Use the **Export Matrix (CSV)** button at the bottom of the grid to export this filtered set."
            )
        else:
            answer_text = (
                f"{prefix}Applied filters to Allocation Matrix: {filters_list_str}.\n\n"
                f"No employees matched these criteria.\n\n"
                f"The Allocation Matrix grid has been updated (showing 0 matches)."
            )
            
        return {
            "answer": answer_text,
            "filters": filters_to_apply
        }

    # Check conversation history for follow-up choice questions
    if history:
        last_assistant_msg = ""
        for msg in reversed(history):
            if msg.get("role") == "assistant":
                last_assistant_msg = msg.get("content", "")
                break
                
        if last_assistant_msg and "Did you mean:" in last_assistant_msg:
            # Extract candidates bullet points
            lines = last_assistant_msg.split("\n")
            candidates = []
            for line in lines:
                if line.strip().startswith("*"):
                    # Strip bullets, asterisks
                    cand = line.replace("*", "").replace("[Local Heuristic Engine Fallback]", "").strip()
                    if cand:
                        candidates.append(cand)
            
            if candidates:
                user_sel = q_lower.strip()
                matched_candidate_strs = []
                
                # Check for "all of them", "both", "every one", "all"
                if any(x in user_sel for x in ["all", "both", "every", "each"]):
                    matched_candidate_strs = candidates
                else:
                    # Check for positional indicators
                    num_map = {
                        "first": 0, "1": 0, "one": 0,
                        "second": 1, "2": 1, "two": 1,
                        "third": 2, "3": 2, "three": 2,
                        "fourth": 3, "4": 3, "four": 3,
                    }
                    for word, idx in num_map.items():
                        if word == user_sel or f"the {word}" in user_sel:
                            if idx < len(candidates):
                                matched_candidate_strs = [candidates[idx]]
                                break
                                
                    if not matched_candidate_strs:
                        # Try fuzzy match directly against candidate strings
                        cand_match, _ = fuzzy_match_ambiguous(user_sel, candidates, lambda x: x)
                        if cand_match:
                            matched_candidate_strs = [cand_match]
                            
                if matched_candidate_strs:
                    results = []
                    for cand_str in matched_candidate_strs:
                        if ":" in cand_str:
                            c_type, c_val = cand_str.split(":", 1)
                            c_type = c_type.strip().lower()
                            c_val = c_val.strip()
                        else:
                            c_type = "unknown"
                            c_val = cand_str.strip()
                            
                        if "location" in c_type:
                            matched_emps = [e for e in employees if e.location.lower() == c_val.lower()]
                            staff_list = [f"**{e.name}** ({e.team}) - {emp_alloc_sums.get(e.id, 0.0):.1f}% utilization" for e in matched_emps]
                            nl = "\n"
                            res = f"**Location: {c_val}** staff:\n" + (nl.join(staff_list) if staff_list else "No planned staff.")
                            results.append(res)
                        elif "team" in c_type:
                            matched_emps = [e for e in employees if e.team.lower() == c_val.lower()]
                            staff_list = [f"**{e.name}** - {emp_alloc_sums.get(e.id, 0.0):.1f}% utilization" for e in matched_emps]
                            nl = "\n"
                            res = f"**Team: {c_val}** members:\n" + (nl.join(staff_list) if staff_list else "No planned staff.")
                            results.append(res)
                        elif "topic" in c_type or "project" in c_type:
                            topic = next((t for t in topics if t.name.lower() == c_val.lower()), None)
                            if topic:
                                allocated_staff = []
                                for emp in employees:
                                    pct = alloc_map.get((emp.id, topic.id), 0.0)
                                    if pct > 0.0:
                                        allocated_staff.append(f"**{emp.name}** ({emp.team}) - {pct}% allocation")
                                nl = "\n"
                                res = f"**Topic: {topic.name}** staff:\n" + (nl.join(allocated_staff) if allocated_staff else "No allocated staff.")
                                results.append(res)
                            else:
                                results.append(f"Could not resolve project '{c_val}'.")
                                
                    nl_double = "\n\n"
                    return {"answer": prefix + f"Here are the details for your selection:\n\n{nl_double.join(results)}"}

    # 1. Try real local LLM (Ollama)
    rag_context = get_planning_rag_context(db, active_scenario.id)
    
    # Format conversational history for LLM
    history_context = ""
    if history:
        history_context = "Conversation History:\n"
        for h in history:
            role_name = "User" if h.get("role") == "user" else "Assistant"
            history_context += f"{role_name}: {h.get('content')}\n"
        history_context += "\n"
        
    system_prompt = (
        "You are the Textron engineering planning AI assistant. You run locally and confidentially on Textron's server.\n"
        "Here are your STRICT rules and guardrails:\n"
        "1. CONFIDENTIALITY: Do not try to access external APIs or search engines. Never reveal user passwords.\n"
        "2. SCOPE: Answer ONLY questions related to resource planning, employees, cost calculation, allocations, and topics/projects "
        "found in the context data. Decline answering unrelated general questions by stating: 'I am a confidential resource planning assistant for Textron. That query is outside my planning scope.'\n"
        "3. FACTUALITY: Base all statistics, costs, and allocations strictly on the provided context. If the data does not contain the answer, "
        "say: 'I cannot find that in the active planning dataset.' Do not make up or assume any values.\n"
        "4. FORMATTING: Return answers in a clear, concise bulleted or tabular format.\n"
        "5. HUMAN-LIKE UNDERSTANDING & CLARIFICATION: Be flexible with spelling, typos, abbreviations, or partial names. If a name, project, or location matches multiple potential entries, do not guess: list the matching candidates and ask the user to clarify between them. Otherwise, if there is a unique best match, present the answer.\n\n"
        f"Context:\n{rag_context}\n\n"
        f"{history_context}"
        f"User Question: {query}\n"
        "Answer:"
    )
    
    ollama_answer = query_local_ollama(system_prompt)
    if ollama_answer:
        return {"answer": ollama_answer}

    q = query.lower().rstrip("?.! ")
    
    # Who is working on/in/at [target]... (Unified Dispatcher)
    match_who = re.search(r"(?:who is working on|who works on|who is allocated to|who is working in|who works in|who is located in|who is on|who works for|employees in|staff in|employees on|staff on)\s+(.+)", q)
    if match_who:
        target = match_who.group(1).strip().replace("project", "").replace("team", "").strip()
        
        # Match against locations
        unique_locs = list(set(e.location for e in employees))
        loc_match, loc_cands = fuzzy_match_ambiguous(target, unique_locs, lambda l: l)
        
        # Match against teams
        unique_teams = list(set(e.team for e in employees))
        team_match, team_cands = fuzzy_match_ambiguous(target, unique_teams, lambda t: t)
        
        # Match against topics
        topic_match, topic_cands = fuzzy_match_ambiguous(target, topics, lambda t: t.name + " " + t.category)
        
        matches = []
        if loc_match: matches.append(("location", loc_match))
        if team_match: matches.append(("team", team_match))
        if topic_match: matches.append(("topic", topic_match))
        
        cands = []
        if loc_cands: cands.extend(("location", c) for c in loc_cands)
        if team_cands: cands.extend(("team", c) for c in team_cands)
        if topic_cands: cands.extend(("topic", c) for c in topic_cands)
        
        if len(matches) == 0 and len(cands) == 0:
            return {"answer": prefix + f"I couldn't find any location, team, or project matching '{target}'."}
            
        if len(cands) > 0 or len(matches) > 1:
            all_options = []
            for m_type, item in matches:
                name = item if m_type == "location" else (item.team if m_type == "team" else item.name)
                all_options.append(f"{m_type.capitalize()}: **{name}**")
            for c_type, item in cands:
                name = item if c_type == "location" else (item if isinstance(item, str) else item.name)
                all_options.append(f"{c_type.capitalize()}: **{name}**")
            c_str = "\n".join(f"* {opt}" for opt in set(all_options))
            return {"answer": prefix + f"I found multiple matches for '{target}'. Did you mean:\n{c_str}\nPlease clarify your query."}
            
        m_type, item = matches[0]
        if m_type == "location":
            matched_emps = [e for e in employees if e.location == item]
            staff_list = [f"**{e.name}** ({e.team}) - {emp_alloc_sums.get(e.id, 0.0):.1f}% utilization" for e in matched_emps]
            nl = "\n"
            return {"answer": prefix + f"Here are the employees working in location **{item}**:\n\n{nl.join(staff_list)}" if staff_list else prefix + f"No employees are currently planned in location **{item}**."}
            
        elif m_type == "team":
            matched_emps = [e for e in employees if e.team == item]
            staff_list = [f"**{e.name}** - {emp_alloc_sums.get(e.id, 0.0):.1f}% utilization" for e in matched_emps]
            nl = "\n"
            return {"answer": prefix + f"Here are the employees working on team **{item}**:\n\n{nl.join(staff_list)}" if staff_list else prefix + f"No employees are currently planned on team **{item}**."}
            
        elif m_type == "topic":
            allocated_staff = []
            for emp in employees:
                pct = alloc_map.get((emp.id, item.id), 0.0)
                if pct > 0.0:
                    allocated_staff.append(f"**{emp.name}** ({emp.team}) - {pct}% allocation")
            nl = "\n"
            return {"answer": prefix + f"Here are the employees working on topic **{item.name}**:\n\n{nl.join(allocated_staff)}" if allocated_staff else prefix + f"Currently, no employees are allocated to **{item.name}**."}

    # What projects are in/of [target]...
    match_what_topics = re.search(r"(?:what projects|what topics|list projects|list topics|projects in|topics in|projects of|topics of|projects on|topics on)\s+(.+)", q)
    if match_what_topics:
        target = match_what_topics.group(1).strip().replace("location", "").replace("category", "").strip()
        
        # Match against locations
        unique_locs = list(set(e.location for e in employees))
        loc_match, loc_cands = fuzzy_match_ambiguous(target, unique_locs, lambda l: l)
        
        # Match against categories
        unique_cats = list(set(t.category for t in topics))
        cat_match, cat_cands = fuzzy_match_ambiguous(target, unique_cats, lambda c: c)
        
        matches = []
        if loc_match: matches.append(("location", loc_match))
        if cat_match: matches.append(("category", cat_match))
        
        cands = []
        if loc_cands: cands.extend(("location", c) for c in loc_cands)
        if cat_cands: cands.extend(("category", c) for c in cat_cands)
        
        if len(matches) == 0 and len(cands) == 0:
            return {"answer": prefix + f"I couldn't find any location or topic category matching '{target}'."}
            
        if len(cands) > 0 or len(matches) > 1:
            all_options = []
            for m_type, item in matches:
                all_options.append(f"{m_type.capitalize()}: **{item}**")
            for c_type, item in cands:
                all_options.append(f"{c_type.capitalize()}: **{item}**")
            c_str = "\n".join(f"* {opt}" for opt in set(all_options))
            return {"answer": prefix + f"I found multiple matches for '{target}'. Did you mean:\n{c_str}\nPlease clarify your query."}
            
        m_type, item = matches[0]
        if m_type == "location":
            loc_emps = [e for e in employees if e.location == item]
            loc_emp_ids = set(e.id for e in loc_emps)
            loc_topics = []
            for t in topics:
                allocated = any(alloc_map.get((emp_id, t.id), 0.0) > 0 for emp_id in loc_emp_ids)
                if allocated:
                    loc_topics.append(t)
                    
            if not loc_topics:
                return {"answer": prefix + f"No topics have allocations from the **{item}** location."}
            t_list = [f"* **{t.name}** ({t.category}) - recovery: ${t.recovery:,.0f}" for t in loc_topics]
            nl = "\n"
            return {"answer": prefix + f"Here are the active topics/projects planned in location **{item}**:\n\n{nl.join(t_list)}"}
            
        elif m_type == "category":
            cat_topics = [t for t in topics if t.category == item]
            if not cat_topics:
                return {"answer": prefix + f"No topics found in category **{item}**."}
            t_list = [f"* **{t.name}** ({t.area}) - recovery: ${t.recovery:,.0f}" for t in cat_topics]
            nl = "\n"
            return {"answer": prefix + f"Here are the topics/projects belonging to category **{item}**:\n\n{nl.join(t_list)}"}

    # What is the total cost of...
    match_cost = re.search(r"(?:total cost of|cost of topic|cost of project|what does)\s+(.+?)(?:\s+cost)?$", q)
    if match_cost:
        target = match_cost.group(1).strip().replace("project", "").strip()
        topic, candidates = fuzzy_match_ambiguous(target, topics, lambda t: t.name)
        if candidates:
            c_str = "\n".join(f"* **{c.name}**" for c in candidates)
            return {"answer": prefix + f"I found multiple topics matching '{target}'. Did you mean:\n{c_str}\nPlease clarify your query."}
        if topic:
            emp_cost = sum(emp.available_hours * emp.hourly_rate * (alloc_map.get((emp.id, topic.id), 0.0) / 100.0) for emp in employees)
            add_int = sum(cost.amount for cost in topic.additional_costs if cost.cost_type == "internal")
            ext_cost = sum(cost.amount for cost in topic.additional_costs if cost.cost_type == "external")
            total = emp_cost + add_int + ext_cost - topic.recovery
            
            return {
                "answer": prefix + f"The total planning cost for **{topic.name}** is **${total:,.2f} USD**.\n\n"
                                   f"* *Internal Staff Cost*: ${emp_cost:,.2f} USD\n"
                                   f"* *Additional Internal Cost*: ${add_int:,.2f} USD\n"
                                   f"* *External Cost*: ${ext_cost:,.2f} USD\n"
                                   f"* *Cost Recovery*: -${topic.recovery:,.2f} USD"
            }

    # Match overloaded employees
    if any(x in q for x in ["overloaded", "over-allocated", "above 100%", "overutilised", "overutilized"]):
        overloaded = []
        for emp in employees:
            tot = emp_alloc_sums.get(emp.id, 0.0)
            if tot > 100.0:
                overloaded.append(f"**{emp.name}** ({emp.team}) - **{tot:.1f}%** total utilization")
        if overloaded:
            nl = "\n"
            return {"answer": prefix + f"The following employees are overloaded (utilization > 100%):\n\n{nl.join(overloaded)}"}
        else:
            return {"answer": prefix + "Excellent! No employees are overloaded in the current active scenario."}

    # Match cost by location
    match_loc = re.search(r"cost of location\s+(.+)", q) or re.search(r"location\s+(.+)\s+cost", q)
    if match_loc:
        loc = match_loc.group(1).strip()
        unique_locs = list(set(e.location for e in employees))
        loc_match, candidates = fuzzy_match_ambiguous(loc, unique_locs, lambda l: l)
        if candidates:
            c_str = "\n".join(f"* **{c}**" for c in candidates)
            return {"answer": prefix + f"I found multiple locations matching '{loc}'. Did you mean:\n{c_str}\nPlease clarify your query."}
        if not loc_match:
            return {"answer": prefix + f"I couldn't find any data for location '{loc}'."}
        
        matched_emps = [e for e in employees if e.location == loc_match]
        loc_cost = 0.0
        for emp in matched_emps:
            for topic in topics:
                pct = alloc_map.get((emp.id, topic.id), 0.0)
                loc_cost += emp.available_hours * emp.hourly_rate * (pct / 100.0)
        return {"answer": prefix + f"The total internal employee cost generated by the **{loc_match}** hub is **${loc_cost:,.2f} USD**."}

    # Match cost by team
    match_team = re.search(r"cost of team\s+(.+)", q) or re.search(r"team\s+(.+)\s+cost", q)
    if match_team:
        team_search = match_team.group(1).strip()
        unique_teams = list(set(e.team for e in employees))
        team_match, candidates = fuzzy_match_ambiguous(team_search, unique_teams, lambda t: t)
        if candidates:
            c_str = "\n".join(f"* **{c}**" for c in candidates)
            return {"answer": prefix + f"I found multiple teams matching '{team_search}'. Did you mean:\n{c_str}\nPlease clarify your query."}
        if not team_match:
            return {"answer": prefix + f"I couldn't find any team matching '{team_search}'."}
            
        team_cost = 0.0
        team_members = [e for e in employees if e.team == team_match]
        for emp in team_members:
            for topic in topics:
                pct = alloc_map.get((emp.id, topic.id), 0.0)
                team_cost += emp.available_hours * emp.hourly_rate * (pct / 100.0)
        return {"answer": prefix + f"The total cost generated by team **{team_match}** is **${team_cost:,.2f} USD** across {len(team_members)} planned members."}

    # CAE / Test filters
    if "cae cost" in q or "cost of cae" in q or "cae topic" in q:
        cae_cost = 0.0
        for emp in employees:
            for topic in topics:
                if "cae" in topic.name.lower() or "cae" in topic.category.lower() or "cae" in emp.department.lower():
                    pct = alloc_map.get((emp.id, topic.id), 0.0)
                    cae_cost += emp.available_hours * emp.hourly_rate * (pct / 100.0)
        return {"answer": prefix + f"The total estimated cost for all **CAE-related** planning activities is **${cae_cost:,.2f} USD**."}

    if "test cost" in q or "cost of test" in q or "test topic" in q:
        test_cost = 0.0
        for emp in employees:
            for topic in topics:
                if "test" in topic.name.lower() or "test" in topic.category.lower() or "test" in emp.department.lower():
                    pct = alloc_map.get((emp.id, topic.id), 0.0)
                    test_cost += emp.available_hours * emp.hourly_rate * (pct / 100.0)
        return {"answer": prefix + f"The total estimated cost for all **Test-related** planning activities is **${test_cost:,.2f} USD**."}

    # Generalized entity + intent detection. The regex handlers above only match
    # a fixed set of phrasings (e.g. "who is working in X"); this catches more
    # natural phrasings ("tell me the names of the people working in Romania",
    # "how many people are on the CAE Germany team", "budget for Fuel Project")
    # by looking for a known location/team/project mentioned anywhere in the
    # sentence, then deciding what to return from the surrounding intent words.
    unique_locs = list(set(e.location for e in employees))
    unique_teams = list(set(e.team for e in employees))

    def find_entity_in_query(text, choices):
        best, best_len = None, 0
        for choice in choices:
            c_lower = choice.lower()
            if c_lower and c_lower in text and len(c_lower) > best_len:
                best, best_len = choice, len(c_lower)
        return best

    loc_hit = find_entity_in_query(q, unique_locs)
    team_hit = None if loc_hit else find_entity_in_query(q, unique_teams)
    topic_hit, topic_hit_candidates = (None, [])
    if not loc_hit and not team_hit:
        topic_hit, topic_hit_candidates = fuzzy_match_ambiguous(q, topics, lambda t: t.name, threshold=0.35)

    if topic_hit_candidates:
        c_str = "\n".join(f"* Topic/Project: **{c.name}**" for c in topic_hit_candidates)
        return {"answer": prefix + f"I found multiple projects matching your question. Did you mean:\n{c_str}\nPlease clarify your query."}

    if loc_hit or team_hit or topic_hit:
        if loc_hit:
            matched_emps = [e for e in employees if e.location == loc_hit]
            entity_label = f"location **{loc_hit}**"
        elif team_hit:
            matched_emps = [e for e in employees if e.team == team_hit]
            entity_label = f"team **{team_hit}**"
        else:
            matched_emps = [e for e in employees if alloc_map.get((e.id, topic_hit.id), 0.0) > 0]
            entity_label = f"topic **{topic_hit.name}**"

        cost_intent = any(w in q for w in ["cost", "budget", "spend", "expense"])
        count_intent = any(w in q for w in ["how many", "count", "number of"])
        people_intent = any(w in q for w in ["name", "names", "who", "people", "employee", "staff", "member"])

        if cost_intent and not people_intent:
            cost_total = sum(
                emp.available_hours * emp.hourly_rate * (alloc_map.get((emp.id, t.id), 0.0) / 100.0)
                for emp in matched_emps for t in topics
            )
            return {"answer": prefix + f"The total cost associated with {entity_label} is **${cost_total:,.2f} USD**."}

        if count_intent and not people_intent:
            return {"answer": prefix + f"There are **{len(matched_emps)}** employees associated with {entity_label}."}

        # Default (and any "names"/"who"/"people" phrasing) -> list the people
        if matched_emps:
            names_list = "\n".join(f"* **{e.name}** ({e.team}) - {emp_alloc_sums.get(e.id, 0.0):.1f}% utilization" for e in matched_emps)
            return {"answer": prefix + f"Here are the people working on {entity_label}:\n\n{names_list}"}
        return {"answer": prefix + f"No employees are currently planned for {entity_label}."}

    # Fallback to search profile (Fuzzy Matcher)
    emp, emp_candidates = fuzzy_match_ambiguous(q, employees, lambda e: e.name)
    topic, topic_candidates = fuzzy_match_ambiguous(q, topics, lambda t: t.name)
    
    if emp_candidates or topic_candidates:
        all_candidates = []
        if emp_candidates:
            all_candidates.extend(f"Employee: **{c.name}**" for c in emp_candidates)
        if topic_candidates:
            all_candidates.extend(f"Topic/Project: **{c.name}**" for c in topic_candidates)
        c_str = "\n".join(all_candidates)
        return {"answer": prefix + f"I found multiple matches for '{query}'. Did you mean:\n{c_str}\nPlease clarify your query."}
        
    if emp:
        tot = emp_alloc_sums.get(emp.id, 0.0)
        cost_contrib = sum(emp.available_hours * emp.hourly_rate * (alloc_map.get((emp.id, t.id), 0.0) / 100.0) for t in topics)
        return {
            "answer": prefix + f"Here is the summary for employee **{emp.name}**:\n"
                               f"* *Team/Location*: {emp.team} / {emp.location}\n"
                               f"* *Hourly Rate*: ${emp.hourly_rate:.2f} USD/hr\n"
                               f"* *Annual Available Hours*: {emp.available_hours:.1f} hrs\n"
                               f"* *Total Utilization*: **{tot:.1f}%**\n"
                               f"* *Total Allocated Cost*: **${cost_contrib:,.2f} USD**"
        }
        
    if topic:
        emp_cost = sum(emp.available_hours * emp.hourly_rate * (alloc_map.get((emp.id, topic.id), 0.0) / 100.0) for emp in employees)
        return {
            "answer": prefix + f"Here is the summary for topic **{topic.name}**:\n"
                               f"* *Category/Area*: {topic.category} / {topic.area}\n"
                               f"* *Internal Staff Cost*: ${emp_cost:,.2f} USD\n"
                               f"* *Recovery/Savings*: -${topic.recovery:,.2f} USD\n"
                               f"* *Description*: {topic.description or 'No description'}"
        }

    # Nothing matched: the question is too vague or references something outside
    # the current dataset. Ask a clarifying question instead of just failing.
    return {
        "answer": prefix + "I want to help, but I'm not sure exactly what you're asking. Could you clarify?\n\n"
                           "* Are you asking about **people/staff**, **costs**, or **projects/topics**?\n"
                           "* Could you mention a specific **team**, **location**, or **project name**?\n\n"
                           "For example: *'Who is working in Romania?'*, *'What is the cost of the Fuel project?'*, or *'List overloaded employees'*."
    }

# Serve static files for frontend SPA
app.mount("/", StaticFiles(directory="static", html=True), name="static")
