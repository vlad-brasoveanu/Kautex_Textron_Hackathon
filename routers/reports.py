import io
import os
import csv
import uuid
import json
import datetime
import difflib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import models
import schemas
from database import get_db
from dependencies import (
    get_current_user,
    require_admin,
    require_master,
    write_system_log,
    get_active_scenario_db,
    build_dashboard_report,
    query_local_ollama,
)

router = APIRouter(prefix="/api", tags=["reports"])

UPLOAD_STORAGE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
os.makedirs(UPLOAD_STORAGE_DIR, exist_ok=True)

class ExportLogPayload(BaseModel):
    report_name: str
    format: str

# AUDIT LOGS ROUTERS
@router.post("/reports/log-export")
def log_report_export(payload: ExportLogPayload, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    write_system_log(
        db,
        username=current_user.username,
        action="Export Report",
        details=f"Exported resource planning report '{payload.report_name}' in {payload.format.upper()} format"
    )
    return {"status": "success"}

@router.get("/admin/logs", response_model=List[schemas.AuditLogResponse])
def get_admin_logs(db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    return db.query(models.SystemLog).order_by(models.SystemLog.timestamp.desc()).all()

@router.delete("/admin/logs")
def clear_admin_logs(db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    count = db.query(models.SystemLog).delete()
    db.commit()
    write_system_log(db, username=master.username, action="Clear Audit Logs", details=f"Permanently cleared {count} audit log entries")
    return {"message": f"Cleared {count} audit log entries"}

# DASHBOARD ROUTERS
@router.get("/reports/dashboard")
def get_dashboard_reports(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    return build_dashboard_report(db, active_scenario)

@router.get("/reports/dashboard/{scenario_id}")
def get_dashboard_report_for_scenario(scenario_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    scenario = db.query(models.Scenario).filter(models.Scenario.id == scenario_id).first()
    if not scenario:
        raise HTTPException(status_code=404, detail="Scenario not found")
    return build_dashboard_report(db, scenario)

# AI MEMO
@router.post("/reports/ai-memo")
def generate_ai_memo(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    report = build_dashboard_report(db, active_scenario)

    top_topics = report["highest_cost_topics"][:3]
    top_depts = sorted(report["cost_by_department"].items(), key=lambda x: x[1], reverse=True)[:3]
    overloaded = report["overloaded_employees"]

    top_topics_str = ", ".join(f"{t['name']} (${t['total_cost']:,.0f})" for t in top_topics) or "none"
    top_depts_str = ", ".join(f"{d} (${c:,.0f})" for d, c in top_depts) or "none"
    overloaded_names_str = ", ".join(e["name"] for e in overloaded) or "none"

    fallback_paragraphs = [
        f"This plan supports {report['total_headcount']} employees at a total annual planning cost of "
        f"${report['total_annual_planning_cost']:,.0f}, combining ${report['total_internal_employee_cost']:,.0f} in internal "
        f"staff cost, ${report['total_additional_internal_cost']:,.0f} in additional internal cost, and "
        f"${report['total_external_cost']:,.0f} in external cost, net of ${report['total_recovery_cost']:,.0f} in recovery.",

        f"The largest cost drivers are {top_topics_str}. By department, {top_depts_str} represent the highest spend.",

        (f"{len(overloaded)} employee(s) are currently allocated above 100% and should be reviewed for rebalancing: "
         f"{overloaded_names_str}." if overloaded else
         "No employees are currently over-allocated - staffing levels across the portfolio are within capacity."),
    ]
    fallback_memo = "\n\n".join(fallback_paragraphs)

    prompt = (
        "You are writing a concise 3-paragraph executive memo for Textron leadership summarizing the current "
        "engineering resource plan. Use ONLY the real data below - do not invent numbers, names, or projects.\n\n"
        f"Scenario: {report['scenario_name']}\n"
        f"Headcount: {report['total_headcount']}\n"
        f"Total annual planning cost: ${report['total_annual_planning_cost']:,.0f}\n"
        f"Highest cost topics: {top_topics_str}\n"
        f"Highest cost departments: {top_depts_str}\n"
        f"Overloaded employees (>100% allocated): {overloaded_names_str}\n\n"
        "Write exactly 3 short paragraphs: (1) overall scope and cost, (2) where the money goes, "
        "(3) risks/recommendations. Plain prose, no headers, no bullet points, no markdown."
    )
    llm_memo = query_local_ollama(prompt)

    return {"memo": llm_memo.strip() if llm_memo else fallback_memo}

# CSV / EXCEL IMPORTER
COST_ROW_KEYWORDS = ["CAD", "Sampling", "Equipment", "PTF", "Testing", "Tooling", "Prototypes",
                      "Supplier", "Services", "Recovery", "Engineering", "Internal", "External", "Other"]

def classify_cost_category(cost_category: str) -> str:
    cat_lower = cost_category.lower()
    if "recovery" in cat_lower:
        return "recovery"
    if cat_lower.startswith("internal") or cat_lower in ("cad", "engineering support", "sampling", "ptf"):
        return "internal"
    return "external"

def parse_uploaded_file_to_rows(contents: bytes, filename: str, content_type: Optional[str]):
    filename = (filename or "").lower()
    rows = []

    is_excel = (filename.endswith(".xlsx") or filename.endswith(".xls")
                or content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"])

    if is_excel:
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                if not any(v is not None for v in row):
                    continue
                row_vals = []
                for val in row:
                    if val is None:
                        row_vals.append("")
                    else:
                        row_vals.append(str(val))
                rows.append(row_vals)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {str(e)}")
        file_type = "excel"
    else:
        try:
            decoded = contents.decode("utf-8")
            csv_file = io.StringIO(decoded)
            reader = csv.reader(csv_file)
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to parse CSV file: {str(e)}")
        file_type = "csv"

    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    return rows, file_type

CANONICAL_FIELD_ALIASES = {
    "employee": ["Employee", "Employee Name", "Name"],
    "team": ["Team"],
    "location": ["Location"],
    "hours": ["Hours / Year", "Available hours/year (Region 100%)", "Hours/Year"],
    "rate": ["Hourly Rate", "Hourly rate"],
    "department": ["Department", "Dept", "Area"],
    "manager": ["Manager"],
    "notes": ["Notes"],
    "status": ["Status"],
}
CANONICAL_FIELD_LABELS = {
    "employee": "Employee", "team": "Team", "location": "Location",
    "hours": "Hours / Year", "rate": "Hourly Rate", "department": "Department",
    "manager": "Manager", "notes": "Notes", "status": "Status",
}

def resolve_header_indices(headers: list, column_mapping: Optional[dict] = None) -> dict:
    def find_idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return -1

    idxs = {field: find_idx(*aliases) for field, aliases in CANONICAL_FIELD_ALIASES.items()}

    for header, field in (column_mapping or {}).items():
        if header in headers and field in idxs:
            idxs[field] = headers.index(header)

    return idxs

def suggest_header_mappings(headers: list, idxs: dict, existing_topic_names: list) -> list:
    matched_idxs = {i for i in idxs.values() if i != -1}
    existing_topic_names_lower = {n.lower() for n in existing_topic_names}
    unmatched_fields = [field for field, i in idxs.items() if i == -1]

    suggestions = []
    for i, h in enumerate(headers):
        if i in matched_idxs or not h:
            continue
        h_lower = h.lower()
        if h_lower in existing_topic_names_lower:
            continue
        if "total" in h_lower or "utilization" in h_lower:
            continue

        best_field, best_score = None, 0.0
        for field in unmatched_fields:
            label_lower = CANONICAL_FIELD_LABELS[field].lower()
            score = difflib.SequenceMatcher(None, h_lower, label_lower).ratio()
            if len(h_lower) >= 2 and (h_lower in label_lower or label_lower in h_lower):
                score += 0.15
            if score > best_score:
                best_field, best_score = field, score

        if best_field and best_score >= 0.55:
            suggestions.append({
                "header": h,
                "suggested_field": best_field,
                "suggested_label": CANONICAL_FIELD_LABELS[best_field],
                "confidence": round(best_score, 2)
            })

    return suggestions

def apply_rows_to_scenario(rows, active_scenario, db: Session, column_mapping: Optional[dict] = None):
    headers = [h.strip() for h in rows[0]]
    idxs = resolve_header_indices(headers, column_mapping)
    emp_idx = idxs["employee"]
    if emp_idx == -1:
        raise HTTPException(status_code=400, detail="Required column 'Employee' not found in CSV headers.")

    team_idx = idxs["team"]
    loc_idx = idxs["location"]
    hours_idx = idxs["hours"]
    rate_idx = idxs["rate"]
    dept_idx = idxs["department"]
    manager_idx = idxs["manager"]
    notes_idx = idxs["notes"]
    status_idx = idxs["status"]

    known_meta_idxs = {i for i in idxs.values() if i != -1}

    topic_cols = []
    for idx, h in enumerate(headers):
        if idx in known_meta_idxs:
            continue
        if "total" in h.lower() or "utilization" in h.lower() or h == "":
            continue
        topic_cols.append((idx, h))

    db_employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id).all()
    emp_ids = [e.id for e in db_employees]
    db.query(models.Allocation).filter(models.Allocation.employee_id.in_(emp_ids)).delete(synchronize_session=False)
    
    additional_cost_ids = [
        ac.id for ac in db.query(models.AdditionalCost)
        .join(models.Topic)
        .filter(models.Topic.scenario_id == active_scenario.id)
        .all()
    ]
    if additional_cost_ids:
        db.query(models.AdditionalCost).filter(models.AdditionalCost.id.in_(additional_cost_ids)).delete(synchronize_session=False)

    existing_emp = {e.name: e for e in db.query(models.Employee).filter(
        models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()}
    existing_top = {t.name: t for t in db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()}

    original_emp_names = set(existing_emp.keys())
    original_top_names = set(existing_top.keys())
    file_topic_names = {name for _, name in topic_cols}
    touched_emp_names = set()

    added_emps = 0
    added_tops = 0
    added_allocs = 0
    added_costs = 0

    def cell(row, idx):
        if idx == -1 or idx >= len(row):
            return ""
        return row[idx].strip()

    last_seen_team = ""
    last_seen_location = ""

    for r_idx in range(1, len(rows)):
        row = rows[r_idx]
        if not row or len(row) <= emp_idx:
            continue

        first_cell = cell(row, emp_idx)
        row_team = cell(row, team_idx)
        row_location = cell(row, loc_idx)

        if row_team:
            last_seen_team = row_team
        if row_location:
            last_seen_location = row_location

        is_cost_row = (not row_team and not row_location and first_cell
                       and any(kw in first_cell for kw in COST_ROW_KEYWORDS))

        if is_cost_row:
            cost_category = first_cell
            cost_type = classify_cost_category(cost_category)

            for col_idx, topic_name in topic_cols:
                if col_idx < len(row) and row[col_idx].strip():
                    val_str = row[col_idx].strip().replace("$", "").replace(",", "").replace("%", "")
                    try:
                        amt = float(val_str)
                        if amt != 0.0:
                            topic = existing_top.get(topic_name)
                            if topic:
                                if cost_type == "recovery":
                                    topic.recovery = abs(amt)
                                else:
                                    db_cost = models.AdditionalCost(
                                        topic_id=topic.id,
                                        cost_type=cost_type,
                                        category=cost_category,
                                        amount=amt,
                                        notes="Imported from CSV"
                                    )
                                    db.add(db_cost)
                                    added_costs += 1
                    except ValueError:
                        pass
            continue

        if not first_cell:
            continue

        emp_name = first_cell
        touched_emp_names.add(emp_name)
        team = row_team or last_seen_team
        location = row_location or last_seen_location
        dept_raw = cell(row, dept_idx)
        manager_raw = cell(row, manager_idx)
        notes_raw = cell(row, notes_idx)
        status_raw = cell(row, status_idx)

        available_hours = None
        hours_raw = cell(row, hours_idx)
        if hours_raw:
            try:
                available_hours = float(hours_raw.replace(",", "").strip())
            except ValueError:
                available_hours = None

        hourly_rate = None
        rate_raw = cell(row, rate_idx)
        if rate_raw:
            try:
                hourly_rate = float(rate_raw.replace("$", "").replace(",", "").strip())
            except ValueError:
                hourly_rate = None

        employee = existing_emp.get(emp_name)
        if not employee:
            dept = dept_raw if dept_raw else ("CAE" if "cae" in team.lower() else "Test" if "test" in team.lower() else "Management")
            employee = models.Employee(
                scenario_id=active_scenario.id,
                name=emp_name,
                team=team or "Unassigned",
                department=dept,
                location=location or "Unassigned",
                available_hours=available_hours if available_hours is not None else 1800.0,
                hourly_rate=hourly_rate if hourly_rate is not None else 50.0,
                status=status_raw or "Active",
                manager=manager_raw or None,
                notes=notes_raw or None
            )
            db.add(employee)
            db.commit()
            db.refresh(employee)
            existing_emp[emp_name] = employee
            added_emps += 1
        else:
            if team:
                employee.team = team
            if location:
                employee.location = location
            if available_hours is not None:
                employee.available_hours = available_hours
            if hourly_rate is not None:
                employee.hourly_rate = hourly_rate
            if dept_raw:
                employee.department = dept_raw
            if manager_raw:
                employee.manager = manager_raw
            if notes_raw:
                employee.notes = notes_raw
            if status_raw:
                employee.status = status_raw
            db.commit()

        for col_idx, topic_name in topic_cols:
            if col_idx < len(row):
                pct_str = row[col_idx].strip().replace("%", "")
                if pct_str:
                    try:
                        pct = float(pct_str)
                        if 0.0 < pct <= 1.0:
                            pct = pct * 100.0
                            
                        if pct > 0.0:
                            topic = existing_top.get(topic_name)
                            if not topic:
                                cat = "Internal Efforts" if "non bookable" in topic_name.lower() or "consulting" in topic_name.lower() else "Internal D-Projects" if "ai" in topic_name.lower() else "Customer Requests"
                                topic = models.Topic(
                                    scenario_id=active_scenario.id,
                                    name=topic_name,
                                    category=cat,
                                    area="CAE" if "cae" in topic_name.lower() else "Test" if "test" in topic_name.lower() else "General",
                                    status="Active"
                                )
                                db.add(topic)
                                db.commit()
                                db.refresh(topic)
                                existing_top[topic_name] = topic
                                added_tops += 1
                                
                            db_alloc = models.Allocation(
                                employee_id=employee.id,
                                topic_id=topic.id,
                                percentage=pct,
                                comment="Imported from CSV"
                            )
                            db.add(db_alloc)
                            added_allocs += 1
                    except ValueError:
                        pass
                        
    archived_emps = 0
    for name in original_emp_names - touched_emp_names:
        emp = existing_emp[name]
        if not emp.is_deleted:
            emp.is_deleted = True
            emp.deleted_at = datetime.datetime.utcnow()
            archived_emps += 1

    archived_tops = 0
    for name in original_top_names - file_topic_names:
        top = existing_top[name]
        if not top.is_deleted:
            top.is_deleted = True
            top.deleted_at = datetime.datetime.utcnow()
            archived_tops += 1

    db.commit()
    return {
        "imported_employees": added_emps,
        "imported_topics": added_tops,
        "imported_allocations": added_allocs,
        "imported_additional_costs": added_costs,
        "archived_employees": archived_emps,
        "archived_topics": archived_tops
    }

@router.post("/import/preview")
async def preview_import_columns(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    contents = await file.read()
    rows, file_type = parse_uploaded_file_to_rows(contents, file.filename, file.content_type)
    headers = [h.strip() for h in rows[0]]
    idxs = resolve_header_indices(headers)

    existing_topics = db.query(models.Topic).filter(
        models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False
    ).all()
    suggestions = suggest_header_mappings(headers, idxs, [t.name for t in existing_topics])

    return {"headers": headers, "suggested_mappings": suggestions}

@router.post("/import/csv")
async def import_csv_data(file: UploadFile = File(...), column_mapping: Optional[str] = Form(None), db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)

    contents = await file.read()
    rows, file_type = parse_uploaded_file_to_rows(contents, file.filename, file.content_type)

    mapping_dict = None
    if column_mapping:
        try:
            mapping_dict = json.loads(column_mapping)
        except (ValueError, TypeError):
            mapping_dict = None

    counts = apply_rows_to_scenario(rows, active_scenario, db, column_mapping=mapping_dict)

    stored_filename = f"{uuid.uuid4().hex}_{os.path.basename(file.filename or 'upload')}"
    with open(os.path.join(UPLOAD_STORAGE_DIR, stored_filename), "wb") as f:
        f.write(contents)

    upload_record = models.UploadHistory(
        scenario_id=active_scenario.id,
        original_filename=file.filename or "upload",
        stored_filename=stored_filename,
        file_type=file_type,
        size_bytes=len(contents),
        uploaded_by=admin.username,
        imported_employees=counts["imported_employees"],
        imported_topics=counts["imported_topics"],
        imported_allocations=counts["imported_allocations"],
        imported_additional_costs=counts["imported_additional_costs"],
        archived_employees=counts["archived_employees"],
        archived_topics=counts["archived_topics"],
        file_content=contents
    )
    db.add(upload_record)
    db.commit()

    archived_note = ""
    if counts["archived_employees"] or counts["archived_topics"]:
        archived_note = (f" Moved {counts['archived_employees']} employee(s) and {counts['archived_topics']} topic(s) "
                          f"to Trash (not present in this file).")

    write_system_log(
        db,
        username=admin.username,
        action="Import CSV",
        details=f"Successfully imported '{file.filename}'. Loaded {counts['imported_employees']} employees, "
                f"{counts['imported_topics']} topics, {counts['imported_allocations']} allocations, {counts['imported_additional_costs']} costs."
                f"{archived_note}"
    )
    return {
        "status": "success",
        "message": f"Successfully parsed and loaded planning sheet into '{active_scenario.name}'",
        **counts
    }

@router.get("/uploads/history", response_model=List[schemas.UploadHistoryResponse])
def get_upload_history(db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    active_scenario = get_active_scenario_db(db)
    return (
        db.query(models.UploadHistory)
        .filter(models.UploadHistory.scenario_id == active_scenario.id)
        .order_by(models.UploadHistory.uploaded_at.desc())
        .all()
    )

@router.delete("/uploads/history/{upload_id}")
def delete_upload_history(upload_id: int, db: Session = Depends(get_db), master: models.User = Depends(require_master)):
    record = db.query(models.UploadHistory).filter(models.UploadHistory.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Upload history record not found")

    stored_path = os.path.join(UPLOAD_STORAGE_DIR, record.stored_filename)
    if os.path.exists(stored_path):
        os.remove(stored_path)

    filename = record.original_filename
    db.delete(record)
    db.commit()
    write_system_log(db, username=master.username, action="Delete Upload History", details=f"Permanently deleted upload history record for '{filename}'")
    return {"message": f"Deleted upload history record for '{filename}'"}

@router.post("/uploads/history/{upload_id}/apply")
def apply_upload_from_history(upload_id: int, db: Session = Depends(get_db), admin: models.User = Depends(require_admin)):
    record = db.query(models.UploadHistory).filter(models.UploadHistory.id == upload_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Upload history record not found")

    if record.file_content is not None:
        contents = record.file_content
    else:
        stored_path = os.path.join(UPLOAD_STORAGE_DIR, record.stored_filename)
        if not os.path.exists(stored_path):
            raise HTTPException(status_code=404, detail="The stored file for this upload is no longer available")
        with open(stored_path, "rb") as f:
            contents = f.read()

    active_scenario = get_active_scenario_db(db)

    content_type = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                     if record.file_type == "excel" else "text/csv")
    rows, _ = parse_uploaded_file_to_rows(contents, record.original_filename, content_type)
    counts = apply_rows_to_scenario(rows, active_scenario, db)

    archived_note = ""
    if counts["archived_employees"] or counts["archived_topics"]:
        archived_note = (f" Moved {counts['archived_employees']} employee(s) and {counts['archived_topics']} topic(s) "
                          f"to Trash (not present in this file).")

    write_system_log(
        db,
        username=admin.username,
        action="Import CSV",
        details=f"Re-applied historical upload '{record.original_filename}' (originally uploaded {record.uploaded_at.strftime('%Y-%m-%d %H:%M')} UTC "
                f"by {record.uploaded_by}) onto '{active_scenario.name}'. Loaded {counts['imported_employees']} employees, "
                f"{counts['imported_topics']} topics, {counts['imported_allocations']} allocations, {counts['imported_additional_costs']} costs."
                f"{archived_note}"
    )
    return {
        "status": "success",
        "message": f"Successfully re-applied '{record.original_filename}' onto '{active_scenario.name}'",
        **counts
    }

@router.get("/export/excel")
def export_excel_data(
    location: Optional[str] = None,
    team: Optional[str] = None,
    department: Optional[str] = None,
    category: Optional[str] = None,
    minRate: Optional[float] = None,
    maxRate: Optional[float] = None,
    manager: Optional[str] = None,
    status: Optional[str] = None,
    topicId: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    active_scenario = get_active_scenario_db(db)

    # Load data
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()
    
    alloc_map = {}
    for a in allocations:
        alloc_map[(a.employee_id, a.topic_id)] = a.percentage

    # Apply filters
    filtered_employees = []
    for emp in employees:
        if location and emp.location != location: continue
        if team and emp.team != team: continue
        if department and emp.department != department: continue
        if manager and emp.manager != manager: continue
        if status and emp.status != status: continue
        if minRate is not None and emp.hourly_rate < minRate: continue
        if maxRate is not None and emp.hourly_rate > maxRate: continue
        if topicId is not None:
            pct = alloc_map.get((emp.id, topicId), 0.0)
            if pct <= 0.0: continue
        filtered_employees.append(emp)
        
    filtered_topics = []
    for t in topics:
        if category and t.category != category: continue
        filtered_topics.append(t)
        
    # Build a styled workbook matching the app's own brand look (dark blue
    # header band, currency/percentage number formats, zebra-striped rows,
    # a distinct fill for the cost/recovery summary rows) rather than a bare
    # openpyxl default. The data reflects whatever filters are currently
    # active in the UI, since filtered_employees/filtered_topics above are
    # already narrowed by the same query params the matrix uses.
    BRAND_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    COST_ROW_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    RECOVERY_ROW_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    THIN_BORDER = Border(*(Side(style="thin", color="CBD5E1") for _ in range(4)))
    TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
    SUBTITLE_FONT = Font(italic=True, size=10, color="E0E7FF")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BOLD_FONT = Font(bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Allocation Matrix"

    num_cols = 5 + len(filtered_topics)
    last_col_letter = get_column_letter(num_cols)

    # Row 1: title banner
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "Textron Digital Engineering - Resource Allocation Matrix"
    title_cell.font = TITLE_FONT
    title_cell.fill = BRAND_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: subtitle - scenario name, export time, and active filters so the
    # sheet is self-documenting about which slice of the data it represents.
    active_filters = []
    if location: active_filters.append(f"Location={location}")
    if team: active_filters.append(f"Team={team}")
    if department: active_filters.append(f"Department={department}")
    if category: active_filters.append(f"Topic Category={category}")
    if manager: active_filters.append(f"Manager={manager}")
    if status: active_filters.append(f"Status={status}")
    if topicId is not None: active_filters.append(f"Active Project Allocation={topicId}")
    if minRate is not None: active_filters.append(f"Min Rate={minRate}")
    if maxRate is not None: active_filters.append(f"Max Rate={maxRate}")
    filters_text = f" | Filters: {', '.join(active_filters)}" if active_filters else " | No filters applied"

    ws.merge_cells(f"A2:{last_col_letter}2")
    subtitle_cell = ws["A2"]
    subtitle_cell.value = f"{active_scenario.name} - Exported {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}{filters_text}"
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.fill = BRAND_FILL
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Row 3: blank spacer
    header_row_idx = 4

    # Row 4: column headers
    headers = ["Employee", "Team", "Location", "Hours/Year", "Hourly Rate"]
    for t in filtered_topics:
        headers.append(t.name)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row_idx].height = 28

    # Employee rows
    row_idx = header_row_idx
    for i, emp in enumerate(filtered_employees):
        row_idx += 1
        values = [emp.name, emp.team, emp.location, emp.available_hours, emp.hourly_rate]
        for t in filtered_topics:
            values.append(alloc_map.get((emp.id, t.id), 0.0))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if i % 2 == 1:
                cell.fill = ZEBRA_FILL
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx == 4:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 5:
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_idx > 5:
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="center")

    # Additional costs
    category_rows = {}
    for t in filtered_topics:
        for ac in t.additional_costs:
            if ac.category not in category_rows:
                category_rows[ac.category] = {}
            category_rows[ac.category][t.id] = category_rows[ac.category].get(t.id, 0.0) + ac.amount

    has_recovery = any(t.recovery and t.recovery != 0 for t in filtered_topics)

    if category_rows or has_recovery:
        row_idx += 1  # blank separator row

    for cat_name, topic_costs in category_rows.items():
        row_idx += 1
        values = [cat_name, "", "", "", ""]
        for t in filtered_topics:
            values.append(topic_costs.get(t.id, ""))
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = COST_ROW_FILL
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx > 5 and val != "":
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    if has_recovery:
        row_idx += 1
        values = ["Recovery", "", "", "", ""]
        for t in filtered_topics:
            values.append(t.recovery if t.recovery else "")
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = RECOVERY_ROW_FILL
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.font = BOLD_FONT
            elif col_idx > 5 and val != "":
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    # Column widths sized to content, and freeze the header row + Employee
    # column so both stay visible while scrolling a large matrix.
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    for col_idx, t in enumerate(filtered_topics, start=6):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(32, len(t.name) // 1.3))
    ws.freeze_panes = f"B{header_row_idx + 1}"

    # Save workbook to memory stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    headers_resp = {
        'Content-Disposition': f'attachment; filename="Allocation_Matrix_{active_scenario.name.replace(" ", "_")}.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)

@router.get("/reports/ai-predictions")
def get_ai_predictions(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    active_scenario = get_active_scenario_db(db)
    employees = db.query(models.Employee).filter(models.Employee.scenario_id == active_scenario.id, models.Employee.is_deleted == False).all()
    topics = db.query(models.Topic).filter(models.Topic.scenario_id == active_scenario.id, models.Topic.is_deleted == False).all()
    allocations = db.query(models.Allocation).join(models.Employee).join(models.Topic).filter(
        models.Employee.scenario_id == active_scenario.id,
        models.Employee.is_deleted == False,
        models.Topic.is_deleted == False
    ).all()
    
    emp_alloc_sums = {}
    for a in allocations:
        emp_alloc_sums[a.employee_id] = emp_alloc_sums.get(a.employee_id, 0.0) + a.percentage
        
    overloaded = [e for e in employees if emp_alloc_sums.get(e.id, 0.0) > 100.0]
    
    predictions = {
        "bottlenecks": [],
        "cost_optimizations": [],
        "reallocations": []
    }
    
    if overloaded:
        for emp in overloaded:
            total_pct = emp_alloc_sums[emp.id]
            predictions["bottlenecks"].append({
                "type": "Resource Overload",
                "severity": "High",
                "description": f"Employee **{emp.name}** is overloaded at **{total_pct:.1f}%** utilization. This predicts burn-out or project milestone delivery delays."
            })
    else:
        predictions["bottlenecks"].append({
            "type": "Resource Staffing",
            "severity": "Low",
            "description": "All employee allocations are within safe bounds (<= 100%). Overall staff utilization risk is low."
        })
        
    alloc_topic_sums = {}
    for a in allocations:
        alloc_topic_sums[a.topic_id] = alloc_topic_sums.get(a.topic_id, 0.0) + a.percentage
        
    for t in topics:
        tot = alloc_topic_sums.get(t.id, 0.0)
        if tot > 0.0 and tot < 50.0:
            predictions["bottlenecks"].append({
                "type": "Underallocation Risk",
                "severity": "Medium",
                "description": f"Topic **{t.name}** has only **{tot:.1f}%** total planned resource allocation. Deliverables may be delayed due to insufficient staff load."
            })

    for t in topics:
        add_ext = sum(ac.amount for ac in t.additional_costs if ac.cost_type == "external")
        emp_cost = 0.0
        for emp in employees:
            alloc_pct = sum(a.percentage for a in allocations if a.employee_id == emp.id and a.topic_id == t.id)
            emp_cost += emp.available_hours * emp.hourly_rate * (alloc_pct / 100.0)
            
        if add_ext > emp_cost and emp_cost > 0:
            savings = add_ext * 0.15
            predictions["cost_optimizations"].append({
                "category": "External Scope Optimization",
                "impact": f"Save ~${savings:,.2f}",
                "description": f"Topic **{t.name}** has high external vendor tooling/costs (${add_ext:,.2f}) compared to internal employee cost (${emp_cost:,.2f}). Suggest sampling internally to cut costs."
            })
            
    if not predictions["cost_optimizations"]:
        predictions["cost_optimizations"].append({
            "category": "Portfolio Savings",
            "impact": "Save ~$12,500.00",
            "description": "Review recovery rates on internal initiatives. Standardizing equipment sampling rates across Germany and Romania hubs is predicted to recover 15% budget."
        })

    if overloaded:
        for emp in overloaded:
            total_pct = emp_alloc_sums[emp.id]
            excess = total_pct - 100.0
            same_team_under = [e for e in employees if e.team == emp.team and emp_alloc_sums.get(e.id, 0.0) < 80.0]
            if same_team_under:
                helper = same_team_under[0]
                predictions["reallocations"].append({
                    "action": "Balance Staff Load",
                    "priority": "High",
                    "description": f"Reallocate **{excess:.1f}%** of load from overloaded **{emp.name}** to **{helper.name}** ({helper.team}) who has available bandwidth (currently at {emp_alloc_sums.get(helper.id, 0.0):.1f}%)."
                })
            else:
                predictions["reallocations"].append({
                    "action": "Hire External Vendor Support",
                    "priority": "Medium",
                    "description": f"Deploy external sampling contract to cover the excess **{excess:.1f}%** workload for **{emp.name}** to mitigate milestone delays."
                })
    else:
        predictions["reallocations"].append({
            "action": "Resource Optimization",
            "priority": "Low",
            "description": "No reallocations needed. Bandwidths are balanced across all planned employees."
        })

    for t in topics:
        topic_total = alloc_topic_sums.get(t.id, 0.0)
        if topic_total <= 0.0:
            continue
        topic_allocs = [a for a in allocations if a.topic_id == t.id]
        top_alloc = max(topic_allocs, key=lambda a: a.percentage, default=None)
        if top_alloc and (top_alloc.percentage / topic_total) >= 0.6:
            emp_name = next((e.name for e in employees if e.id == top_alloc.employee_id), "Unknown")
            predictions["bottlenecks"].append({
                "type": "Single-Point-of-Failure Risk",
                "severity": "High",
                "description": f"**{emp_name}** alone carries **{(top_alloc.percentage / topic_total) * 100:.0f}%** of the total effort planned on **{t.name}**. Losing this person (leave, attrition) would stall the topic with no backup coverage."
            })

    location_costs = {}
    location_headcounts = {}
    for emp in employees:
        emp_cost = emp.available_hours * emp.hourly_rate * (emp_alloc_sums.get(emp.id, 0.0) / 100.0)
        location_costs[emp.location] = location_costs.get(emp.location, 0.0) + emp_cost
        location_headcounts[emp.location] = location_headcounts.get(emp.location, 0) + 1
    location_avgs = {loc: location_costs[loc] / location_headcounts[loc] for loc in location_costs if location_headcounts[loc] > 0}
    if location_avgs:
        portfolio_avg = sum(location_avgs.values()) / len(location_avgs)
        for loc, avg in location_avgs.items():
            if portfolio_avg > 0 and avg > portfolio_avg * 1.5:
                predictions["cost_optimizations"].append({
                    "category": "Location Cost Outlier",
                    "impact": f"${avg - portfolio_avg:,.0f}/employee above portfolio average",
                    "description": f"**{loc}** averages **${avg:,.0f}** in annual cost per employee, over 1.5x the portfolio average of ${portfolio_avg:,.0f}. Review whether this location's rates or allocation levels are proportionate to its role in the plan."
                })

    for t in topics:
        if not t.justification or len(t.justification.strip()) < 15:
            predictions["bottlenecks"].append({
                "type": "Governance Gap: Missing Business Justification",
                "severity": "Low",
                "description": f"Topic **{t.name}** has no (or a very thin) documented business justification. Undocumented spend is harder to defend in budget reviews - add a justification before the next planning cycle."
            })

    predictions = enrich_predictions_with_llm(predictions, employees, topics, allocations, emp_alloc_sums, alloc_topic_sums)
    return predictions

def enrich_predictions_with_llm(predictions: dict, employees: list, topics: list, allocations: list, emp_alloc_sums: dict, alloc_topic_sums: dict) -> dict:
    top_util = sorted(
        [(e.name, e.team, emp_alloc_sums.get(e.id, 0.0)) for e in employees],
        key=lambda x: x[2], reverse=True
    )[:8]
    under_topics = sorted(
        [(t.name, t.category, alloc_topic_sums.get(t.id, 0.0)) for t in topics],
        key=lambda x: x[2]
    )[:8]

    data_summary = (
        "Employee utilization (name, team, total allocation %):\n" +
        "\n".join(f"- {n} ({team}): {pct:.1f}%" for n, team, pct in top_util) +
        "\n\nTopic total allocation (name, category, total allocation %):\n" +
        "\n".join(f"- {n} ({cat}): {pct:.1f}%" for n, cat, pct in under_topics)
    )

    prompt = (
        "You are a resource planning analyst for Textron. Based ONLY on the real portfolio data below, "
        "suggest up to 2 additional bottleneck risks, up to 2 additional cost optimizations, and up to 2 "
        "additional staff reallocation actions that are NOT obvious duplicates of simple overload/underload "
        "statements. Be specific and reference the real names/topics given.\n\n"
        f"Portfolio Data:\n{data_summary}\n\n"
        "Respond with STRICT JSON only, no prose, matching exactly this shape:\n"
        '{"bottlenecks": [{"type": "...", "severity": "High|Medium|Low", "description": "..."}], '
        '"cost_optimizations": [{"category": "...", "impact": "...", "description": "..."}], '
        '"reallocations": [{"action": "...", "priority": "High|Medium|Low", "description": "..."}]}'
    )

    raw = query_local_ollama(prompt)
    if not raw:
        return predictions

    try:
        start = raw.index("{")
        end = raw.rindex("}") + 1
        llm_data = json.loads(raw[start:end])
    except (ValueError, json.JSONDecodeError):
        return predictions

    for key, required_keys in (
        ("bottlenecks", {"type", "severity", "description"}),
        ("cost_optimizations", {"category", "impact", "description"}),
        ("reallocations", {"action", "priority", "description"}),
    ):
        items = llm_data.get(key)
        if not isinstance(items, list):
            continue
        for item in items[:2]:
            if isinstance(item, dict) and required_keys.issubset(item.keys()):
                predictions[key].append(item)

    return predictions
