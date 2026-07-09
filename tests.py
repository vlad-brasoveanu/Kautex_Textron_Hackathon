import pytest
import io
import csv
import json
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from database import Base
import models
from main import app, get_db, hash_password, create_access_token
from fastapi.testclient import TestClient

# Create sqlite in-memory database for testing
SQLALCHEMY_DATABASE_URL = "sqlite:///./test_planning.db"
engine = create_engine(SQLALCHEMY_DATABASE_URL, connect_args={"check_same_thread": False})
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Override database dependency in FastAPI app
def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()

app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)

# Helper headers - real signed JWTs (same mechanism the cookie carries),
# sent as a Bearer header since the test client doesn't go through a
# browser login flow to pick up the HttpOnly cookie.
master_headers = {"Authorization": f"Bearer {create_access_token('master', 'master_admin')}"}
admin_headers = {"Authorization": f"Bearer {create_access_token('admin', 'admin')}"}
user_headers = {"Authorization": f"Bearer {create_access_token('user', 'user')}"}
invalid_headers = {"Authorization": "Bearer not-a-valid-token"}

@pytest.fixture(autouse=True)
def setup_database():
    # The TestClient is shared across the whole module, so a session cookie
    # set by an earlier test's real /api/auth/login call would otherwise
    # leak into later tests (e.g. ones checking unauthenticated requests).
    client.cookies.clear()

    # Setup database schema before each test run
    Base.metadata.create_all(bind=engine)
    db = TestingSessionLocal()
    
    # Clean previous records
    db.query(models.Allocation).delete()
    db.query(models.AdditionalCost).delete()
    db.query(models.Employee).delete()
    db.query(models.Topic).delete()
    db.query(models.Scenario).delete()
    db.query(models.User).delete()
    db.commit()
    
    # Seed default users
    master_user = models.User(
        username="master",
        password_hash=hash_password("master123"),
        name="Master Manager",
        role="master_admin"
    )
    admin_user = models.User(
        username="admin",
        password_hash=hash_password("admin123"),
        name="Admin Director",
        role="admin"
    )
    regular_user = models.User(
        username="user",
        password_hash=hash_password("user123"),
        name="Staff Analyst",
        role="user"
    )
    db.add(master_user)
    db.add(admin_user)
    db.add(regular_user)
    
    # Insert Test scenario
    scenario = models.Scenario(name="Test Scenario", description="Automation validation", is_active=True)
    db.add(scenario)
    db.commit()
    db.refresh(scenario)
    
    yield db
    
    # Tear down database
    Base.metadata.drop_all(bind=engine)

def test_login_auth_endpoint():
    # Test valid admin login
    response = client.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
    assert response.status_code == 200
    assert response.json()["role"] == "admin"
    assert response.json()["access_token"]  # a real signed JWT, not a fixed string
    assert response.cookies.get("session_token")

    # Test valid user login
    response = client.post("/api/auth/login", json={"username": "user", "password": "user123"})
    assert response.status_code == 200
    assert response.json()["role"] == "user"
    
    # Test invalid login
    response = client.post("/api/auth/login", json={"username": "admin", "password": "wrongpassword"})
    assert response.status_code == 401

def test_security_auth_guards(setup_database):
    # Test unauthorized access (no headers)
    response = client.get("/api/scenarios")
    assert response.status_code in (401, 403)

    # Test invalid token access
    response = client.get("/api/scenarios", headers=invalid_headers)
    assert response.status_code == 401

    # Test authorized read access
    response = client.get("/api/scenarios", headers=user_headers)
    assert response.status_code == 200

    # Test user (read-only) trying to create a scenario
    response = client.post("/api/scenarios", json={"name": "New Scen"}, headers=user_headers)
    assert response.status_code == 403  # Forbidden write for normal user

    # Test admin (read-write) creating a scenario
    response = client.post("/api/scenarios", json={"name": "Admin Scen"}, headers=admin_headers)
    assert response.status_code == 200

def test_scenario_crud(setup_database):
    response = client.get("/api/scenarios", headers=user_headers)
    assert response.status_code == 200
    assert len(response.json()) == 1
    assert response.json()[0]["name"] == "Test Scenario"

def test_cost_calculation_engine(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()
    
    emp = models.Employee(
        scenario_id=scenario.id,
        name="Tester A",
        team="CAE Germany",
        department="CAE",
        location="Germany",
        available_hours=1000.0,
        hourly_rate=100.0,
        status="Active"
    )
    db.add(emp)
    
    topic = models.Topic(
        scenario_id=scenario.id,
        name="AI Testing Initiative",
        category="Internal Efforts",
        area="CAE",
        recovery=1000.0
    )
    db.add(topic)
    db.commit()
    db.refresh(emp)
    db.refresh(topic)
    
    alloc = models.Allocation(
        employee_id=emp.id,
        topic_id=topic.id,
        percentage=50.0,
        comment="Planned half-time on AI testing."
    )
    db.add(alloc)
    
    add_cost_int = models.AdditionalCost(
        topic_id=topic.id,
        cost_type="internal",
        category="CAD",
        amount=2000.0
    )
    add_cost_ext = models.AdditionalCost(
        topic_id=topic.id,
        cost_type="external",
        category="Tooling",
        amount=5000.0
    )
    db.add(add_cost_int)
    db.add(add_cost_ext)
    db.commit()
    
    response = client.get("/api/reports/dashboard", headers=user_headers)
    assert response.status_code == 200
    data = response.json()
    
    assert data["total_internal_employee_cost"] == 50000.0
    assert data["total_additional_internal_cost"] == 2000.0
    assert data["total_external_cost"] == 5000.0
    assert data["total_recovery_cost"] == 1000.0
    assert data["total_annual_planning_cost"] == 56000.0

def test_simulation_dashboard_report_for_arbitrary_scenario(setup_database):
    db = setup_database
    base_scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(
        scenario_id=base_scenario.id, name="Sim Employee", team="Sim Team",
        department="CAE", location="Germany", available_hours=1000.0, hourly_rate=100.0
    )
    topic = models.Topic(scenario_id=base_scenario.id, name="Sim Topic", category="Internal Efforts")
    db.add(emp)
    db.add(topic)
    db.commit()
    db.refresh(emp)
    db.refresh(topic)
    db.add(models.Allocation(employee_id=emp.id, topic_id=topic.id, percentage=50.0))
    db.commit()

    # Clone the base scenario into a simulation sandbox (this is what
    # "Start a Simulation" does) - clone_scenario auto-activates the clone.
    response = client.post(
        f"/api/scenarios/{base_scenario.id}/clone",
        json={"new_name": "Sim Sandbox", "new_description": "[Simulation Sandbox] Cloned from Test Scenario"},
        headers=admin_headers
    )
    assert response.status_code == 200
    sim_scenario_id = response.json()["id"]

    # The dashboard report for an arbitrary (non-active) scenario works even
    # though the sandbox is now the active one, proving Simulation's Compare
    # feature can pull both sides without switching the active scenario back
    # and forth.
    response = client.get(f"/api/reports/dashboard/{base_scenario.id}", headers=admin_headers)
    assert response.status_code == 200
    base_report = response.json()
    assert base_report["scenario_name"] == "Test Scenario"
    assert base_report["total_headcount"] == 1

    response = client.get(f"/api/reports/dashboard/{sim_scenario_id}", headers=admin_headers)
    assert response.status_code == 200
    sim_report = response.json()
    assert sim_report["scenario_name"] == "Sim Sandbox"
    assert sim_report["total_headcount"] == 1
    assert sim_report["cost_by_team"] == base_report["cost_by_team"]

    # Nonexistent scenario -> 404
    response = client.get("/api/reports/dashboard/999999", headers=admin_headers)
    assert response.status_code == 404

    # Simulate a what-if edit on the now-active sandbox scenario (a "Change
    # Hourly Rate" quick action) and confirm it's reflected in that
    # scenario's report but not the original's.
    sim_emp_resp = client.get("/api/employees", headers=admin_headers)
    sim_emp = next(e for e in sim_emp_resp.json() if e["name"] == "Sim Employee")
    response = client.put(
        f"/api/employees/{sim_emp['id']}",
        json={**sim_emp, "hourly_rate": 150.0},
        headers=admin_headers
    )
    assert response.status_code == 200

    sim_report_after = client.get(f"/api/reports/dashboard/{sim_scenario_id}", headers=admin_headers).json()
    base_report_after = client.get(f"/api/reports/dashboard/{base_scenario.id}", headers=admin_headers).json()
    assert sim_report_after["total_internal_employee_cost"] > base_report_after["total_internal_employee_cost"]

    # "Apply to Real Planning Version" = switch the active scenario back to
    # the original baseline, reusing the existing active-scenario endpoint.
    response = client.post(f"/api/scenarios/active/{base_scenario.id}", headers=admin_headers)
    assert response.status_code == 200
    active_resp = client.get("/api/scenarios/active", headers=admin_headers)
    assert active_resp.json()["id"] == base_scenario.id


def test_simulation_sandbox_stays_isolated_until_applied(setup_database):
    """The current Simulation tab design: cloning with activate=False must
    NEVER touch the app-wide active scenario, and all Quick Actions target
    the sandbox exclusively through scenario-scoped endpoints - so the real
    active scenario's data (and everything every other page reads) is
    completely unaffected until an explicit "Apply" switches the pointer."""
    db = setup_database
    base_scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(
        scenario_id=base_scenario.id, name="Real Employee", team="Real Team",
        department="CAE", location="Germany", available_hours=1000.0, hourly_rate=100.0
    )
    db.add(emp)
    db.commit()

    # Clone WITHOUT activating - this is what "Start Simulation" does now.
    response = client.post(
        f"/api/scenarios/{base_scenario.id}/clone",
        json={"new_name": "Isolated Sandbox", "new_description": "[Simulation Sandbox] Cloned from Test Scenario", "activate": False},
        headers=admin_headers
    )
    assert response.status_code == 200
    sandbox = response.json()
    assert sandbox["is_active"] is False

    # The real active scenario is completely untouched.
    active_resp = client.get("/api/scenarios/active", headers=admin_headers)
    assert active_resp.json()["id"] == base_scenario.id
    assert active_resp.json()["name"] == "Test Scenario"

    # The global (active-scenario-bound) employee list is unaffected by the
    # sandbox's existence.
    global_emps = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in global_emps} == {"Real Employee"}

    # A regular user cannot see the sandbox's scoped data (Simulation is admin-only).
    response = client.get(f"/api/scenarios/{sandbox['id']}/employees", headers=user_headers)
    assert response.status_code == 403

    # Scenario-scoped read shows the sandbox's own (cloned) roster.
    sandbox_emps = client.get(f"/api/scenarios/{sandbox['id']}/employees", headers=admin_headers).json()
    assert {e["name"] for e in sandbox_emps} == {"Real Employee"}

    # Scenario-scoped create (the "Add New Employee" Quick Action) adds to
    # the sandbox only - never touches the real scenario's roster.
    response = client.post(
        f"/api/scenarios/{sandbox['id']}/employees",
        json={"name": "Sandbox-Only Hire", "team": "Real Team", "department": "CAE", "location": "Germany"},
        headers=admin_headers
    )
    assert response.status_code == 200
    sandbox_emps = client.get(f"/api/scenarios/{sandbox['id']}/employees", headers=admin_headers).json()
    assert {e["name"] for e in sandbox_emps} == {"Real Employee", "Sandbox-Only Hire"}

    global_emps = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in global_emps} == {"Real Employee"}  # still unaffected

    # Scenario-scoped topic create likewise only affects the sandbox.
    response = client.post(
        f"/api/scenarios/{sandbox['id']}/topics",
        json={"name": "Sandbox-Only Topic", "category": "Internal Efforts"},
        headers=admin_headers
    )
    assert response.status_code == 200
    sandbox_topics = client.get(f"/api/scenarios/{sandbox['id']}/topics", headers=admin_headers).json()
    assert {t["name"] for t in sandbox_topics} == {"Sandbox-Only Topic"}
    assert client.get("/api/topics", headers=admin_headers).json() == []

    # Existing PUT/allocation endpoints are scenario-agnostic by row ID, so
    # they can edit the sandbox's employee directly without any special
    # scenario-scoped variant.
    hire = next(e for e in sandbox_emps if e["name"] == "Sandbox-Only Hire")
    response = client.put(
        f"/api/employees/{hire['id']}",
        json={**hire, "hourly_rate": 999.0},
        headers=admin_headers
    )
    assert response.status_code == 200
    sandbox_emps_after = client.get(f"/api/scenarios/{sandbox['id']}/employees", headers=admin_headers).json()
    assert next(e for e in sandbox_emps_after if e["name"] == "Sandbox-Only Hire")["hourly_rate"] == 999.0

    # Applying the simulation switches the active pointer - and only now
    # does the sandbox's data become visible through the global endpoints.
    response = client.post(f"/api/scenarios/active/{sandbox['id']}", headers=admin_headers)
    assert response.status_code == 200
    global_emps = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in global_emps} == {"Real Employee", "Sandbox-Only Hire"}


def test_scenario_lifecycle_audit_logging(setup_database):
    db = setup_database
    base_scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    response = client.post("/api/scenarios", json={"name": "Brand New Version", "description": "test"}, headers=admin_headers)
    assert response.status_code == 200
    new_id = response.json()["id"]

    response = client.post(f"/api/scenarios/active/{base_scenario.id}", headers=admin_headers)
    assert response.status_code == 200

    response = client.post(
        f"/api/scenarios/{base_scenario.id}/clone",
        json={"new_name": "Cloned Version", "activate": False},
        headers=admin_headers
    )
    assert response.status_code == 200
    cloned_id = response.json()["id"]

    response = client.delete(f"/api/scenarios/{new_id}", headers=admin_headers)
    assert response.status_code == 200

    logs = client.get("/api/admin/logs", headers=admin_headers).json()
    actions = [l["action"] for l in logs]
    assert "Create Scenario" in actions
    assert "Switch Active Scenario" in actions
    assert "Clone Scenario" in actions
    assert "Delete Scenario" in actions

    clone_log = next(l for l in logs if l["action"] == "Clone Scenario")
    assert "Cloned Version" in clone_log["details"]
    assert "sandbox, not activated" in clone_log["details"]


def test_local_ai_query_assistant(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()
    
    emp = models.Employee(
        scenario_id=scenario.id,
        name="Jane Doe",
        team="Test Bonn",
        department="Test",
        location="Germany",
        available_hours=1600.0,
        hourly_rate=120.0,
        status="Active"
    )
    db.add(emp)
    topic = models.Topic(
        scenario_id=scenario.id,
        name="Validation Protocol Fuel",
        category="Customer Requests",
        recovery=0.0
    )
    db.add(topic)
    db.commit()
    db.refresh(emp)
    db.refresh(topic)
    
    alloc = models.Allocation(employee_id=emp.id, topic_id=topic.id, percentage=120.0)
    db.add(alloc)
    db.commit()
    
    response = client.post("/api/ai/query", json={"query": "List overloaded employees"}, headers=user_headers)
    assert response.status_code == 200
    assert "Jane Doe" in response.json()["answer"]
    assert "120.0%" in response.json()["answer"]
    
    response = client.post("/api/ai/query", json={"query": "Who is working on Fuel?"}, headers=user_headers)
    assert response.status_code == 200
    assert "Jane Doe" in response.json()["answer"]
    assert "120.0%" in response.json()["answer"]

    response = client.post("/api/ai/query", json={"query": "who is working in Germany?"}, headers=user_headers)
    assert response.status_code == 200
    assert "Jane Doe" in response.json()["answer"]

    # Natural-language phrasing that doesn't match any of the fixed regex patterns
    response = client.post("/api/ai/query", json={"query": "tell me what are the names of the people working in Germany"}, headers=user_headers)
    assert response.status_code == 200
    assert "Jane Doe" in response.json()["answer"]

    response = client.post("/api/ai/query", json={"query": "how many people are on team Test Bonn"}, headers=user_headers)
    assert response.status_code == 200
    assert "1" in response.json()["answer"]

    response = client.post("/api/ai/query", json={"query": "Invalid question that contains topic cost"}, headers=user_headers)
    assert response.status_code == 200
    assert "not sure exactly what you're asking" in response.json()["answer"]

    # Test out-of-scope guardrail block
    response = client.post("/api/ai/query", json={"query": "What is the capital of France?"}, headers=user_headers)
    assert response.status_code == 200
    assert "guardrails" in response.json()["answer"] or "outside my planning scope" in response.json()["answer"]

def test_ai_chat_executes_allocation_changes(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(
        scenario_id=scenario.id, name="Marcus Reyes", team="CAE Germany",
        department="CAE", location="Germany", hourly_rate=90.0
    )
    topic_fuel = models.Topic(scenario_id=scenario.id, name="Fuel Project", category="Customer Requests")
    topic_agentic = models.Topic(scenario_id=scenario.id, name="Agentic AI", category="Internal Efforts")
    db.add_all([emp, topic_fuel, topic_agentic])
    db.commit()
    db.refresh(emp)
    db.refresh(topic_fuel)
    db.refresh(topic_agentic)

    alloc = models.Allocation(employee_id=emp.id, topic_id=topic_fuel.id, percentage=40.0)
    db.add(alloc)
    db.commit()

    # A regular user cannot execute an allocation change, only get told to ask an admin
    response = client.post("/api/ai/query", json={"query": "set Marcus Reyes's allocation on Fuel Project to 60%"}, headers=user_headers)
    assert response.status_code == 200
    assert "only Admins can execute" in response.json()["answer"]
    assert not response.json().get("action_executed")

    # An admin can set an allocation directly via chat
    response = client.post("/api/ai/query", json={"query": "set Marcus Reyes's allocation on Fuel Project to 60%"}, headers=admin_headers)
    assert response.status_code == 200
    data = response.json()
    assert data.get("action_executed") is True
    assert "60.0%" in data["answer"]

    alloc_resp = client.get("/api/allocations", headers=admin_headers)
    match = next(a for a in alloc_resp.json() if a["employee_id"] == emp.id and a["topic_id"] == topic_fuel.id)
    assert match["percentage"] == 60.0

    # An admin can move allocation between two topics via chat
    response = client.post(
        "/api/ai/query",
        json={"query": "move 20% of Marcus Reyes's Fuel Project time to Agentic AI"},
        headers=admin_headers
    )
    assert response.status_code == 200
    data = response.json()
    assert data.get("action_executed") is True

    alloc_resp = client.get("/api/allocations", headers=admin_headers).json()
    fuel_alloc = next(a for a in alloc_resp if a["employee_id"] == emp.id and a["topic_id"] == topic_fuel.id)
    agentic_alloc = next(a for a in alloc_resp if a["employee_id"] == emp.id and a["topic_id"] == topic_agentic.id)
    assert fuel_alloc["percentage"] == 40.0
    assert agentic_alloc["percentage"] == 20.0

    # Audit log records the AI-triggered change
    log_resp = client.get("/api/admin/logs", headers=admin_headers)
    assert any(l["action"] == "AI Chat Allocation Change" for l in log_resp.json())


def test_ai_what_if_copilot_stays_sandboxed(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(
        scenario_id=scenario.id, name="Priya Desai", team="CAE Romania",
        department="CAE", location="Romania", hourly_rate=70.0
    )
    topic_fuel = models.Topic(scenario_id=scenario.id, name="Fuel Project", category="Customer Requests")
    topic_agentic = models.Topic(scenario_id=scenario.id, name="Agentic AI", category="Internal Efforts")
    db.add_all([emp, topic_fuel, topic_agentic])
    db.commit()
    db.refresh(emp)
    db.refresh(topic_fuel)

    alloc = models.Allocation(employee_id=emp.id, topic_id=topic_fuel.id, percentage=40.0)
    db.add(alloc)
    db.commit()

    # A regular user cannot run a what-if simulation, only get told to ask an admin
    response = client.post(
        "/api/ai/query",
        json={"query": "what if we set Priya Desai's allocation on Fuel Project to 80%"},
        headers=user_headers
    )
    assert response.status_code == 200
    assert "only Admins can run" in response.json()["answer"]
    assert "simulation" not in response.json()

    # An admin's what-if request creates an invisible sandbox and reports a
    # delta, but never touches the real active scenario's data.
    response = client.post(
        "/api/ai/query",
        json={"query": "what if we set Priya Desai's allocation on Fuel Project to 80%"},
        headers=admin_headers
    )
    assert response.status_code == 200
    data = response.json()
    assert "simulation" in data
    sandbox_id = data["simulation"]["scenario_id"]
    assert sandbox_id != scenario.id

    # The real scenario's allocation is untouched.
    alloc_resp = client.get("/api/allocations", headers=admin_headers).json()
    real_alloc = next(a for a in alloc_resp if a["employee_id"] == emp.id and a["topic_id"] == topic_fuel.id)
    assert real_alloc["percentage"] == 40.0

    # The sandbox itself has the simulated change and is tagged so the
    # existing "Clean Up Sandboxes" flow can find it.
    sandbox_emps = client.get(f"/api/scenarios/{sandbox_id}/employees", headers=admin_headers).json()
    sandbox_emp = next(e for e in sandbox_emps if e["name"] == "Priya Desai")
    sandbox_allocs = client.get(f"/api/scenarios/{sandbox_id}/allocations", headers=admin_headers).json()
    sandbox_alloc = next(a for a in sandbox_allocs if a["employee_id"] == sandbox_emp["id"])
    assert sandbox_alloc["percentage"] == 80.0

    scenarios = client.get("/api/scenarios", headers=admin_headers).json()
    sandbox_scenario = next(s for s in scenarios if s["id"] == sandbox_id)
    assert sandbox_scenario["is_active"] is False
    assert "simulation sandbox" in sandbox_scenario["description"].lower()

    active_resp = client.get("/api/scenarios/active", headers=admin_headers)
    assert active_resp.json()["id"] == scenario.id

    # An unparseable what-if request gets a helpful message, no sandbox created.
    scenarios_before = len(client.get("/api/scenarios", headers=admin_headers).json())
    response = client.post("/api/ai/query", json={"query": "what if everything was different"}, headers=admin_headers)
    assert response.status_code == 200
    assert "simulation" not in response.json()
    scenarios_after = len(client.get("/api/scenarios", headers=admin_headers).json())
    assert scenarios_after == scenarios_before


def test_ai_predictions_endpoint(setup_database):
    response = client.get("/api/reports/ai-predictions", headers=user_headers)
    assert response.status_code == 200
    data = response.json()
    assert "bottlenecks" in data
    assert "cost_optimizations" in data
    assert "reallocations" in data


def test_ai_predictions_enriched_by_llm(setup_database, monkeypatch):
    # No LLM reachable -> heuristic-only predictions, unaffected by the enrichment layer.
    import main
    monkeypatch.setattr(main, "query_local_ollama", lambda prompt: None)
    response = client.get("/api/reports/ai-predictions", headers=user_headers)
    assert response.status_code == 200
    baseline_bottleneck_count = len(response.json()["bottlenecks"])

    # A real portfolio-data-fed LLM response should be parsed and appended.
    llm_json = (
        '{"bottlenecks": [{"type": "Skill Concentration Risk", "severity": "Medium", '
        '"description": "Only one CAE engineer covers thermal simulation work."}], '
        '"cost_optimizations": [], "reallocations": []}'
    )
    monkeypatch.setattr(main, "query_local_ollama", lambda prompt: f"Here is my analysis:\n{llm_json}\nHope this helps!")
    response = client.get("/api/reports/ai-predictions", headers=user_headers)
    assert response.status_code == 200
    data = response.json()
    assert len(data["bottlenecks"]) == baseline_bottleneck_count + 1
    assert any(b["type"] == "Skill Concentration Risk" for b in data["bottlenecks"])

    # Malformed LLM output must not break the endpoint - it should just fall back.
    monkeypatch.setattr(main, "query_local_ollama", lambda prompt: "not valid json at all")
    response = client.get("/api/reports/ai-predictions", headers=user_headers)
    assert response.status_code == 200
    assert len(response.json()["bottlenecks"]) == baseline_bottleneck_count

def test_local_ai_conversation_memory(setup_database):
    history = [
        {"role": "user", "content": "who is working in romania?"},
        {"role": "assistant", "content": "[Local Heuristic Engine Fallback] I found multiple matches for 'romania'. Did you mean:\n* Location: Romania\n* Team: Test Romania\n* Team: CAE Romania\nPlease clarify your query."}
    ]
    response = client.post("/api/ai/query", json={"query": "all of them", "history": history}, headers=user_headers)
    assert response.status_code == 200
    assert "Location: Romania" in response.json()["answer"]
    assert "Team: Test Romania" in response.json()["answer"]

def test_audit_logging_scenarios(setup_database):
    # 1. Test Registration
    reg_payload = {"username": "newaudituser", "password": "securepass123", "name": "Audit User", "role": "user"}
    response = client.post("/api/auth/register", json=reg_payload)
    assert response.status_code == 200
    assert response.json()["username"] == "newaudituser"

    # 2. Test Report Export Logging
    export_payload = {"report_name": "Test Excel Export", "format": "CSV"}
    response = client.post("/api/reports/log-export", json=export_payload, headers=user_headers)
    assert response.status_code == 200
    assert response.json()["status"] == "success"

    # 3. Test Admin Logs Retrieval
    response = client.get("/api/admin/logs", headers=admin_headers)
    assert response.status_code == 200
    logs = response.json()
    assert len(logs) > 0
    # The registration and export log events should be present in the audit log list
    actions = [l["action"] for l in logs]
    assert "Registration" in actions
    assert "Export Report" in actions

    # 4. General user cannot read logs
    response = client.get("/api/admin/logs", headers=user_headers)
    assert response.status_code == 403

def test_hierarchical_role_management(setup_database):
    # 1. Master admin can create an admin
    payload = {"username": "new_admin", "password": "password123", "name": "New Admin Name", "role": "admin"}
    response = client.post("/api/users", json=payload, headers=master_headers)
    assert response.status_code == 200
    assert response.json()["username"] == "new_admin"
    assert response.json()["role"] == "admin"

    # 2. Admin cannot create another admin
    payload_bad = {"username": "bad_admin", "password": "password123", "name": "Bad Admin Name", "role": "admin"}
    response = client.post("/api/users", json=payload_bad, headers=admin_headers)
    assert response.status_code == 403

    # 3. Admin can create a standard user
    payload_ok = {"username": "new_user_by_admin", "password": "password123", "name": "User Name", "role": "user"}
    response = client.post("/api/users", json=payload_ok, headers=admin_headers)
    assert response.status_code == 200

    # 4. Standard user cannot create anyone
    response = client.post("/api/users", json=payload_ok, headers=user_headers)
    assert response.status_code == 403

    # 5. Cannot delete master admin (id 1)
    response = client.delete("/api/users/1", headers=master_headers)
    assert response.status_code == 400
    assert "protected" in response.json()["detail"]

def test_edit_user_role_hierarchy(setup_database):
    # Seed an admin and a user account to edit
    client.post("/api/users", json={"username": "edit_admin", "password": "password123", "name": "Edit Admin", "role": "admin"}, headers=master_headers)
    client.post("/api/users", json={"username": "edit_user", "password": "password123", "name": "Edit User", "role": "user"}, headers=master_headers)
    admin_id = next(u["id"] for u in client.get("/api/users", headers=master_headers).json() if u["username"] == "edit_admin")
    user_id = next(u["id"] for u in client.get("/api/users", headers=master_headers).json() if u["username"] == "edit_user")

    # Admin can edit a User's profile details, but cannot change roles
    response = client.put(f"/api/users/{user_id}", json={"name": "Edited By Admin", "department": "Test Dept"}, headers=admin_headers)
    assert response.status_code == 200
    assert response.json()["name"] == "Edited By Admin"
    assert response.json()["department"] == "Test Dept"

    response = client.put(f"/api/users/{user_id}", json={"role": "admin"}, headers=admin_headers)
    assert response.status_code == 403

    # Admin cannot edit the Master Admin's account at all
    response = client.put("/api/users/1", json={"name": "Hacked Name"}, headers=admin_headers)
    assert response.status_code == 403

    # Master can edit anyone, including changing an Admin's role down to User
    response = client.put(f"/api/users/{admin_id}", json={"role": "user"}, headers=master_headers)
    assert response.status_code == 200
    assert response.json()["role"] == "user"

    # Master cannot change their own Master Admin role
    response = client.put("/api/users/1", json={"role": "admin"}, headers=master_headers)
    assert response.status_code == 400

    # Master can edit their own profile details and reset a password
    response = client.put("/api/users/1", json={"name": "Master Manager Updated"}, headers=master_headers)
    assert response.status_code == 200
    assert response.json()["name"] == "Master Manager Updated"

    response = client.put(f"/api/users/{user_id}", json={"password": "newpassword456"}, headers=master_headers)
    assert response.status_code == 200
    login_resp = client.post("/api/auth/login", json={"username": "edit_user", "password": "newpassword456"})
    assert login_resp.status_code == 200


def test_master_only_purge_actions(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(scenario_id=scenario.id, name="Purge Employee", team="Test Team", department="Test", location="Romania", hourly_rate=50.0)
    db.add(emp)
    db.commit()
    db.refresh(emp)

    # Cannot permanently delete a live (non-trashed) employee
    response = client.delete(f"/api/employees/{emp.id}/permanent", headers=master_headers)
    assert response.status_code == 400

    client.delete(f"/api/employees/{emp.id}", headers=admin_headers)

    # An admin cannot permanently delete from Trash - only Master can
    response = client.delete(f"/api/employees/{emp.id}/permanent", headers=admin_headers)
    assert response.status_code == 403

    response = client.delete(f"/api/employees/{emp.id}/permanent", headers=master_headers)
    assert response.status_code == 200

    trash_resp = client.get("/api/trash", headers=master_headers)
    assert not any(e["name"] == "Purge Employee" for e in trash_resp.json()["employees"])

    # Empty Trash bulk-purges everything remaining
    topic = models.Topic(scenario_id=scenario.id, name="Purge Topic", category="Internal Efforts")
    db.add(topic)
    db.commit()
    client.delete(f"/api/topics/{topic.id}", headers=admin_headers)

    response = client.delete("/api/trash", headers=admin_headers)
    assert response.status_code == 403

    response = client.delete("/api/trash", headers=master_headers)
    assert response.status_code == 200
    trash_resp = client.get("/api/trash", headers=master_headers).json()
    assert trash_resp["employees"] == []
    assert trash_resp["topics"] == []

    # Clear Audit Logs (Master only)
    response = client.delete("/api/admin/logs", headers=admin_headers)
    assert response.status_code == 403
    response = client.delete("/api/admin/logs", headers=master_headers)
    assert response.status_code == 200
    logs = client.get("/api/admin/logs", headers=master_headers).json()
    # Only the "Clear Audit Logs" entry itself should remain
    assert len(logs) == 1
    assert logs[0]["action"] == "Clear Audit Logs"

    # Delete Upload History record (Master only)
    csv_data = [["Employee", "Team", "Location", "Hours/Year", "Hourly Rate"], ["Purge Upload User", "Test Team", "Romania", "1800", "60.0"]]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    client.post("/api/import/csv", files={"file": ("purge_test.csv", output.getvalue(), "text/csv")}, headers=admin_headers)
    history = client.get("/api/uploads/history", headers=admin_headers).json()
    record_id = history[0]["id"]

    response = client.delete(f"/api/uploads/history/{record_id}", headers=admin_headers)
    assert response.status_code == 403
    response = client.delete(f"/api/uploads/history/{record_id}", headers=master_headers)
    assert response.status_code == 200
    history = client.get("/api/uploads/history", headers=admin_headers).json()
    assert not any(h["id"] == record_id for h in history)


def test_create_user_optional_profile_fields(setup_database):
    # Optional fields (email, department, position, supervisor) can be supplied...
    payload = {
        "username": "full_profile_user", "password": "password123", "name": "Full Profile",
        "role": "user", "email": "full.profile@textron.com", "department": "CAE Engineering",
        "position": "Senior Engineer", "supervisor": "Jane Boss"
    }
    response = client.post("/api/users", json=payload, headers=admin_headers)
    assert response.status_code == 200
    data = response.json()
    assert data["email"] == "full.profile@textron.com"
    assert data["department"] == "CAE Engineering"
    assert data["position"] == "Senior Engineer"
    assert data["supervisor"] == "Jane Boss"

    # ...and are optional, defaulting to null when omitted
    payload_minimal = {"username": "minimal_profile_user", "password": "password123", "name": "Minimal Profile", "role": "user"}
    response = client.post("/api/users", json=payload_minimal, headers=admin_headers)
    assert response.status_code == 200
    data = response.json()
    assert data["email"] is None
    assert data["department"] is None
    assert data["position"] is None
    assert data["supervisor"] is None

    # 6. Admin cannot delete another admin (id 2)
    response = client.delete("/api/users/2", headers=admin_headers)
    assert response.status_code == 400  # Self-deletion check triggers first because admin_headers is user 2

def test_csv_importer(setup_database):
    csv_data = [
        ["Employee", "Team", "Location", "Hours/Year", "Hourly Rate", "Initiative A"],
        ["GV Test User", "Test Team", "Romania", "1800", "60.0", "50%"],
        ["", "", "", "", "", ""],
        ["Tooling", "", "", "", "", "12000.0"]
    ]
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(csv_data)
    csv_content = output.getvalue()
    
    file_payload = {"file": ("planning_grid.csv", csv_content, "text/csv")}
    response = client.post("/api/import/csv", files=file_payload, headers=admin_headers)
    
    assert response.status_code == 200
    res_data = response.json()
    assert res_data["status"] == "success"
    assert res_data["imported_employees"] == 1
    assert res_data["imported_topics"] == 1
    assert res_data["imported_allocations"] == 1
    assert res_data["imported_additional_costs"] == 1

def test_ai_assisted_column_mapping(setup_database):
    # "Hrly Rate" and "Loc" are mistyped/abbreviated versions of known meta
    # columns - the preview endpoint should suggest remapping them instead of
    # letting them silently become bogus new topic columns.
    csv_data = [
        ["Employee", "Team", "Loc", "Hours/Year", "Hrly Rate", "Fuel Project"],
        ["Mapping User", "Test Team", "Romania", "1800", "70.0", "40%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    csv_content = output.getvalue()

    response = client.post(
        "/api/import/preview",
        files={"file": ("mapping_test.csv", csv_content, "text/csv")},
        headers=admin_headers
    )
    assert response.status_code == 200
    data = response.json()
    assert data["headers"] == ["Employee", "Team", "Loc", "Hours/Year", "Hrly Rate", "Fuel Project"]

    suggestion_headers = {s["header"]: s for s in data["suggested_mappings"]}
    assert "Loc" in suggestion_headers
    assert suggestion_headers["Loc"]["suggested_field"] == "location"
    assert "Hrly Rate" in suggestion_headers
    assert suggestion_headers["Hrly Rate"]["suggested_field"] == "rate"
    # A real topic column should never be suggested as a meta-field remap
    assert "Fuel Project" not in suggestion_headers

    # A regular user cannot preview an import
    response = client.post(
        "/api/import/preview",
        files={"file": ("mapping_test.csv", csv_content, "text/csv")},
        headers=user_headers
    )
    assert response.status_code == 403

    # Without confirming the mapping, "Loc" and "Hrly Rate" are treated as
    # unrecognized columns (i.e. phantom topic columns), not Location/Rate.
    response = client.post(
        "/api/import/csv",
        files={"file": ("mapping_test.csv", csv_content, "text/csv")},
        headers=admin_headers
    )
    assert response.status_code == 200
    emp_resp = client.get("/api/employees", headers=admin_headers)
    mapped_emp = next(e for e in emp_resp.json() if e["name"] == "Mapping User")
    assert mapped_emp["location"] == "Unassigned"
    assert mapped_emp["hourly_rate"] == 50.0  # default, since "Hrly Rate" wasn't recognized

    # Now confirm the suggested mapping and re-import - the columns should
    # resolve correctly this time.
    column_mapping = json.dumps({"Loc": "location", "Hrly Rate": "rate"})
    response = client.post(
        "/api/import/csv",
        files={"file": ("mapping_test.csv", csv_content, "text/csv")},
        data={"column_mapping": column_mapping},
        headers=admin_headers
    )
    assert response.status_code == 200
    emp_resp = client.get("/api/employees", headers=admin_headers)
    mapped_emp = next(e for e in emp_resp.json() if e["name"] == "Mapping User")
    assert mapped_emp["location"] == "Romania"
    assert mapped_emp["hourly_rate"] == 70.0


def test_upload_history_recorded_and_reapplied(setup_database):
    # 1. Upload a file - it should show up in Upload History
    csv_data = [
        ["Employee", "Team", "Location", "Hours/Year", "Hourly Rate", "Initiative A"],
        ["History User", "Test Team", "Romania", "1800", "60.0", "50%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    file_payload = {"file": ("history_test.csv", output.getvalue(), "text/csv")}
    response = client.post("/api/import/csv", files=file_payload, headers=admin_headers)
    assert response.status_code == 200

    response = client.get("/api/uploads/history", headers=admin_headers)
    assert response.status_code == 200
    history = response.json()
    assert len(history) == 1
    record = history[0]
    assert record["original_filename"] == "history_test.csv"
    assert record["file_type"] == "csv"
    assert record["uploaded_by"] == "admin"
    assert record["imported_employees"] == 1

    # 2. A standard user cannot view or apply upload history
    response = client.get("/api/uploads/history", headers=user_headers)
    assert response.status_code == 403
    response = client.post(f"/api/uploads/history/{record['id']}/apply", headers=user_headers)
    assert response.status_code == 403

    # 3. Delete the employee that came from the upload, then re-apply the
    # historical file - it should be recreated, proving the stored file was
    # actually re-parsed and re-applied, not just relabeled.
    emp_resp = client.get("/api/employees", headers=admin_headers)
    hist_emp = next(e for e in emp_resp.json() if e["name"] == "History User")
    response = client.delete(f"/api/employees/{hist_emp['id']}", headers=admin_headers)
    assert response.status_code == 200

    emp_resp = client.get("/api/employees", headers=admin_headers)
    assert not any(e["name"] == "History User" for e in emp_resp.json())

    response = client.post(f"/api/uploads/history/{record['id']}/apply", headers=admin_headers)
    assert response.status_code == 200
    assert response.json()["status"] == "success"

    emp_resp = client.get("/api/employees", headers=admin_headers)
    assert any(e["name"] == "History User" for e in emp_resp.json())

def test_csv_importer_missing_columns_keeps_existing_values(setup_database):
    # First upload establishes full employee data (team, location, hours, rate)
    full_csv = [
        ["Employee", "Team", "Location", "Hours/Year", "Hourly Rate", "Initiative A"],
        ["Partial Col User", "CAE Team", "Germany", "1800", "75.0", "40%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(full_csv)
    client.post("/api/import/csv", files={"file": ("full.csv", output.getvalue(), "text/csv")}, headers=admin_headers)

    # Second upload only carries Employee + a new topic column - Team/Location/Hours/Rate
    # are missing entirely and must NOT be wiped out on the existing employee.
    partial_csv = [
        ["Employee", "Initiative B"],
        ["Partial Col User", "20%"],
    ]
    output2 = io.StringIO()
    csv.writer(output2).writerows(partial_csv)
    response = client.post("/api/import/csv", files={"file": ("partial.csv", output2.getvalue(), "text/csv")}, headers=admin_headers)

    assert response.status_code == 200
    res_data = response.json()
    assert res_data["imported_employees"] == 0  # updated existing, not created
    assert res_data["imported_topics"] == 1  # "Initiative B" is a brand-new column/topic

    emp_resp = client.get("/api/employees", headers=admin_headers)
    employee = next(e for e in emp_resp.json() if e["name"] == "Partial Col User")
    assert employee["team"] == "CAE Team"
    assert employee["location"] == "Germany"
    assert employee["available_hours"] == 1800.0
    assert employee["hourly_rate"] == 75.0

def test_csv_importer_syncs_to_new_template(setup_database):
    # Upload Template A: two employees, one topic column.
    template_a = [
        ["Employee", "Team", "Location", "Hourly Rate", "Legacy Topic"],
        ["Alice Old", "Old Team", "Germany", "50", "40%"],
        ["Bob Old", "Old Team", "Germany", "55", "60%"],
    ]
    output_a = io.StringIO()
    csv.writer(output_a).writerows(template_a)
    response = client.post("/api/import/csv", files={"file": ("template_a.csv", output_a.getvalue(), "text/csv")}, headers=admin_headers)
    assert response.status_code == 200
    data_a = response.json()
    assert data_a["imported_employees"] == 2
    assert data_a["imported_topics"] == 1
    assert data_a["archived_employees"] == 0
    assert data_a["archived_topics"] == 0

    emp_resp = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in emp_resp} == {"Alice Old", "Bob Old"}
    top_resp = client.get("/api/topics", headers=admin_headers).json()
    assert {t["name"] for t in top_resp} == {"Legacy Topic"}

    # Upload Template B: a completely different roster and topic set - only
    # "Alice Old" carries over, "Bob Old" and "Legacy Topic" are gone, and two
    # brand-new employees plus a brand-new topic column show up.
    template_b = [
        ["Employee", "Team", "Location", "Hourly Rate", "New Topic"],
        ["Alice Old", "New Team", "Romania", "65", "25%"],
        ["Carol New", "New Team", "Romania", "45", "35%"],
        ["Dave New", "New Team", "Romania", "48", "50%"],
    ]
    output_b = io.StringIO()
    csv.writer(output_b).writerows(template_b)
    response = client.post("/api/import/csv", files={"file": ("template_b.csv", output_b.getvalue(), "text/csv")}, headers=admin_headers)
    assert response.status_code == 200
    data_b = response.json()
    assert data_b["imported_employees"] == 2  # Carol New, Dave New
    assert data_b["imported_topics"] == 1  # New Topic
    assert data_b["archived_employees"] == 1  # Bob Old
    assert data_b["archived_topics"] == 1  # Legacy Topic

    emp_resp = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in emp_resp} == {"Alice Old", "Carol New", "Dave New"}
    alice = next(e for e in emp_resp if e["name"] == "Alice Old")
    assert alice["team"] == "New Team"
    assert alice["location"] == "Romania"
    assert alice["hourly_rate"] == 65.0

    top_resp = client.get("/api/topics", headers=admin_headers).json()
    assert {t["name"] for t in top_resp} == {"New Topic"}

    # Bob Old and Legacy Topic are recoverable from Trash, not gone forever.
    trash = client.get("/api/trash", headers=admin_headers).json()
    assert any(e["name"] == "Bob Old" for e in trash["employees"])
    assert any(t["name"] == "Legacy Topic" for t in trash["topics"])


def test_csv_importer_team_first_column_order(setup_database):
    # Columns are matched by header name, not position - Team can lead
    # instead of Employee, and each employee still carries its own team
    # value on every row (a flat, non-grouped layout).
    csv_data = [
        ["Team", "Employee", "Location", "Hourly Rate", "Initiative A"],
        ["Team One", "Alice", "Germany", "50", "20%"],
        ["Team One", "Bob", "Germany", "55", "40%"],
        ["Team Two", "Carol", "Romania", "45", "60%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    response = client.post("/api/import/csv", files={"file": ("team_first.csv", output.getvalue(), "text/csv")}, headers=admin_headers)
    assert response.status_code == 200
    assert response.json()["imported_employees"] == 3

    emp_resp = client.get("/api/employees", headers=admin_headers).json()
    alice = next(e for e in emp_resp if e["name"] == "Alice")
    bob = next(e for e in emp_resp if e["name"] == "Bob")
    carol = next(e for e in emp_resp if e["name"] == "Carol")
    assert alice["team"] == "Team One" and alice["location"] == "Germany"
    assert bob["team"] == "Team One" and bob["hourly_rate"] == 55.0
    assert carol["team"] == "Team Two" and carol["location"] == "Romania"


def test_csv_importer_grouped_team_forward_fill(setup_database):
    # Some exports group employees under a team header row (e.g. a merged
    # cell in Excel) and leave Team/Location blank on the employee rows
    # underneath it. Those blanks should forward-fill from the most recent
    # group header instead of falling back to "Unassigned".
    csv_data = [
        ["Team", "Employee", "Location", "Hourly Rate", "Initiative A"],
        ["Team One", "", "Germany", "", ""],
        ["", "Alice Grouped", "", "50", "20%"],
        ["", "Bob Grouped", "", "55", "40%"],
        ["Team Two", "", "Romania", "", ""],
        ["", "Carol Grouped", "", "45", "60%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    response = client.post("/api/import/csv", files={"file": ("grouped.csv", output.getvalue(), "text/csv")}, headers=admin_headers)
    assert response.status_code == 200
    data = response.json()
    # Only the 3 named employee rows are imported - the 2 pure group-header
    # rows (team name only, no employee name) are not created as employees.
    assert data["imported_employees"] == 3

    emp_resp = client.get("/api/employees", headers=admin_headers).json()
    assert {e["name"] for e in emp_resp} == {"Alice Grouped", "Bob Grouped", "Carol Grouped"}

    alice = next(e for e in emp_resp if e["name"] == "Alice Grouped")
    bob = next(e for e in emp_resp if e["name"] == "Bob Grouped")
    carol = next(e for e in emp_resp if e["name"] == "Carol Grouped")
    assert alice["team"] == "Team One" and alice["location"] == "Germany"
    assert bob["team"] == "Team One" and bob["location"] == "Germany"
    assert carol["team"] == "Team Two" and carol["location"] == "Romania"


def test_csv_importer_new_manager_column(setup_database):
    csv_data = [
        ["Employee", "Team", "Location", "Manager", "Initiative A"],
        ["Managed User", "Test Team", "Romania", "Jane Boss", "30%"],
    ]
    output = io.StringIO()
    csv.writer(output).writerows(csv_data)
    response = client.post("/api/import/csv", files={"file": ("managers.csv", output.getvalue(), "text/csv")}, headers=admin_headers)

    assert response.status_code == 200
    emp_resp = client.get("/api/employees", headers=admin_headers)
    employee = next(e for e in emp_resp.json() if e["name"] == "Managed User")
    assert employee["manager"] == "Jane Boss"

def test_excel_importer(setup_database):
    # Construct an Excel sheet in memory using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Employee", "Team", "Location", "Hours/Year", "Hourly Rate", "Initiative Excel"])
    ws.append(["Excel Employee", "Excel Team", "USA", 2000, 85.5, 0.5])
    ws.append(["CAD", "", "", "", "", 15000])
    
    output = io.BytesIO()
    wb.save(output)
    excel_content = output.getvalue()
    
    file_payload = {"file": ("planning_grid.xlsx", excel_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    response = client.post("/api/import/csv", files=file_payload, headers=admin_headers)
    
    assert response.status_code == 200
    res_data = response.json()
    assert res_data["status"] == "success"
    assert res_data["imported_employees"] == 1
    assert res_data["imported_topics"] == 1
    assert res_data["imported_allocations"] == 1
    assert res_data["imported_additional_costs"] == 1

def test_scenario_backup_restore(setup_database):
    # 1. Add some mock data to scenario 1, including a topic with recovery/
    # description and an additional cost - these are the fields that a prior
    # backup/restore bug silently dropped (wrong AdditionalCost field names,
    # and only name/category were preserved on topics).
    emp_payload = {"name": "Backup Employee", "team": "QA", "department": "Engineering", "location": "Romania", "available_hours": 1600, "hourly_rate": 45.0}
    response = client.post("/api/employees", json=emp_payload, headers=admin_headers)
    assert response.status_code == 200

    topic_payload = {"name": "Backup Topic", "category": "Testing", "description": "Full regression suite", "recovery": 2500.0}
    response = client.post("/api/topics", json=topic_payload, headers=admin_headers)
    assert response.status_code == 200
    topic_id = response.json()["id"]

    cost_payload = {"cost_type": "external", "category": "Tooling", "amount": 7500.0, "notes": "Rig rental"}
    response = client.post(f"/api/topics/{topic_id}/costs", json=cost_payload, headers=admin_headers)
    assert response.status_code == 200

    # Get backup of scenario 1
    response = client.get("/api/scenarios/1/backup", headers=admin_headers)
    assert response.status_code == 200
    backup_data = response.json()
    assert backup_data["name"] == "Test Scenario"
    assert len(backup_data["employees"]) > 0
    assert len(backup_data["topics"]) > 0

    backed_up_topic = next(t for t in backup_data["topics"] if t["name"] == "Backup Topic")
    assert backed_up_topic["description"] == "Full regression suite"
    assert backed_up_topic["recovery"] == 2500.0

    backed_up_cost = next(c for c in backup_data["additional_costs"] if c["topic_name"] == "Backup Topic")
    assert backed_up_cost["category"] == "Tooling"
    assert backed_up_cost["amount"] == 7500.0
    assert backed_up_cost["notes"] == "Rig rental"

    # 2. Restore scenario 1 from backup_data
    restore_payload = {
        "name": "Restored Scenario",
        "description": "Successfully restored scenario",
        "employees": backup_data["employees"],
        "topics": backup_data["topics"],
        "allocations": backup_data["allocations"],
        "additional_costs": backup_data["additional_costs"]
    }

    response = client.post("/api/scenarios/1/restore", json=restore_payload, headers=admin_headers)
    assert response.status_code == 200
    assert response.json()["status"] == "success"

    # Verify scenario was updated
    response = client.get("/api/scenarios", headers=admin_headers)
    assert response.status_code == 200
    active_scenario = next(s for s in response.json() if s["is_active"])
    assert active_scenario["name"] == "Restored Scenario"
    assert active_scenario["description"] == "Successfully restored scenario"

    # Verify the restored topic and additional cost kept their full data, not
    # just name/category.
    response = client.get("/api/topics", headers=admin_headers)
    assert response.status_code == 200
    restored_topic = next(t for t in response.json() if t["name"] == "Backup Topic")
    assert restored_topic["description"] == "Full regression suite"
    assert restored_topic["recovery"] == 2500.0
    restored_cost = next(c for c in restored_topic["additional_costs"] if c["category"] == "Tooling")
    assert restored_cost["amount"] == 7500.0
    assert restored_cost["notes"] == "Rig rental"

def test_export_excel_endpoint(setup_database):
    # Verify authentication required
    response = client.get("/api/export/excel")
    assert response.status_code == 401
    
    # Query with authentication headers
    response = client.get("/api/export/excel?location=Romania", headers=user_headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Read the streamed bytes using openpyxl
    file_bytes = io.BytesIO(response.content)
    wb = openpyxl.load_workbook(file_bytes)
    assert "Allocation Matrix" in wb.sheetnames
    
    ws = wb["Allocation Matrix"]
    # Row 1 is a branded title banner, row 2 a subtitle (scenario name/export
    # time/active filters), row 3 a blank spacer, row 4 the column headers.
    assert "Resource Allocation Matrix" in ws["A1"].value
    assert "Location=Romania" in ws["A2"].value

    headers = [cell.value for cell in ws[4]]
    assert "Employee" in headers
    assert "Team" in headers
    assert "Location" in headers

    # Check that location is Romania - row 4 is header, row 5 is first employee
    if ws.max_row > 4:
        loc_idx = headers.index("Location") + 1
        assert ws.cell(row=5, column=loc_idx).value == "Romania"


def test_soft_delete_trash_and_restore(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp = models.Employee(
        scenario_id=scenario.id, name="Trash Employee", team="CAE Germany",
        department="CAE", location="Germany", hourly_rate=80.0
    )
    topic = models.Topic(
        scenario_id=scenario.id, name="Trash Topic", category="Internal Efforts"
    )
    db.add(emp)
    db.add(topic)
    db.commit()
    db.refresh(emp)
    db.refresh(topic)

    # Soft-delete both
    response = client.delete(f"/api/employees/{emp.id}", headers=admin_headers)
    assert response.status_code == 200
    assert response.json()["message"] == "Employee moved to Trash"

    response = client.delete(f"/api/topics/{topic.id}", headers=admin_headers)
    assert response.status_code == 200

    # They disappear from the live lists
    emp_resp = client.get("/api/employees", headers=admin_headers)
    assert not any(e["name"] == "Trash Employee" for e in emp_resp.json())
    top_resp = client.get("/api/topics", headers=admin_headers)
    assert not any(t["name"] == "Trash Topic" for t in top_resp.json())

    # A regular user cannot see the trash
    response = client.get("/api/trash", headers=user_headers)
    assert response.status_code == 403

    # An admin can see both deleted records in the trash
    response = client.get("/api/trash", headers=admin_headers)
    assert response.status_code == 200
    trash = response.json()
    assert any(e["name"] == "Trash Employee" for e in trash["employees"])
    assert any(t["name"] == "Trash Topic" for t in trash["topics"])

    # Restoring brings them back into the live lists
    response = client.post(f"/api/employees/{emp.id}/restore", headers=admin_headers)
    assert response.status_code == 200
    response = client.post(f"/api/topics/{topic.id}/restore", headers=admin_headers)
    assert response.status_code == 200

    emp_resp = client.get("/api/employees", headers=admin_headers)
    assert any(e["name"] == "Trash Employee" for e in emp_resp.json())
    top_resp = client.get("/api/topics", headers=admin_headers)
    assert any(t["name"] == "Trash Topic" for t in top_resp.json())

    # Trash is empty again
    response = client.get("/api/trash", headers=admin_headers)
    trash = response.json()
    assert not any(e["name"] == "Trash Employee" for e in trash["employees"])
    assert not any(t["name"] == "Trash Topic" for t in trash["topics"])


def test_bulk_edit_employees(setup_database):
    db = setup_database
    scenario = db.query(models.Scenario).filter(models.Scenario.is_active == True).first()

    emp1 = models.Employee(
        scenario_id=scenario.id, name="Bulk One", team="Old Team",
        department="CAE", location="Romania", hourly_rate=50.0
    )
    emp2 = models.Employee(
        scenario_id=scenario.id, name="Bulk Two", team="Old Team",
        department="CAE", location="Romania", hourly_rate=100.0
    )
    db.add_all([emp1, emp2])
    db.commit()
    db.refresh(emp1)
    db.refresh(emp2)

    # A regular user cannot bulk edit
    response = client.patch("/api/employees/bulk", json={"employee_ids": [emp1.id], "team": "New Team"}, headers=user_headers)
    assert response.status_code == 403

    # Set team + adjust rate by +10% for both employees
    response = client.patch("/api/employees/bulk", json={
        "employee_ids": [emp1.id, emp2.id],
        "team": "New Team",
        "hourly_rate_adjust_pct": 10.0
    }, headers=admin_headers)
    assert response.status_code == 200
    updated = response.json()
    assert len(updated) == 2
    for u in updated:
        assert u["team"] == "New Team"
    rates = {u["id"]: u["hourly_rate"] for u in updated}
    assert rates[emp1.id] == 55.0
    assert rates[emp2.id] == 110.0

    # Set (not adjust) hourly rate for one employee
    response = client.patch("/api/employees/bulk", json={
        "employee_ids": [emp1.id],
        "hourly_rate_set": 75.0
    }, headers=admin_headers)
    assert response.status_code == 200
    assert response.json()[0]["hourly_rate"] == 75.0

    # A soft-deleted employee is not touched by a bulk edit
    client.delete(f"/api/employees/{emp2.id}", headers=admin_headers)
    response = client.patch("/api/employees/bulk", json={
        "employee_ids": [emp2.id],
        "team": "Should Not Apply"
    }, headers=admin_headers)
    assert response.status_code == 404

